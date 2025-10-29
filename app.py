from datetime import datetime
from functools import wraps
import os
import re
import uuid
import json
import logging
import sqlite3
import concurrent.futures
import threading
import tempfile
import time
from werkzeug.utils import secure_filename
from flask import Flask, session, request, redirect, url_for, render_template, flash, jsonify

try:
    import pandas as pd
except ImportError:
    pd = None
    print("警告: pandas未安装，部分功能将不可用")

# 可选模糊匹配库，若未安装回退到 difflib（返回 0-100）
try:
    from rapidfuzz import fuzz as _rf_fuzz
except Exception:
    import difflib
    def _rf_fuzz(a, b):
        return int(difflib.SequenceMatcher(None, str(a or ''), str(b or '')).ratio() * 100)

# 日志与上传目录
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# 应用常量定义
MAX_UPLOAD_SIZE = 100 * 1024 * 1024  # 100 MB
UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), 'uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# 兼容/占位项
_AEVAL_AVAILABLE = False
Interpreter = None

# Flask app（若文件已定义 app，此处不覆盖）
# 确保静态分析能看到 app 的定义：若全局已有 app 则使用，否则创建一个 Flask 实例
if 'app' not in globals() or globals().get('app') is None:
    app = Flask(__name__)
app.config.setdefault('UPLOAD_FOLDER', UPLOAD_FOLDER)

# 添加：会话密钥与上传限制（请在生产环境通过环境变量设置 SECRET_KEY）
import secrets
app.config['SECRET_KEY'] = secrets.token_hex(16)
app.config.setdefault('MAX_CONTENT_LENGTH', MAX_UPLOAD_SIZE)

# 数据库路径与简单 get_db 实现
DB_PATH = os.path.join(os.path.dirname(__file__), 'zhiguan.db')

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

_get_conn = get_db

def _locate_temp_file(temp_id: str):
    """在 UPLOAD_FOLDER 中按前缀查找上传的临时文件路径"""
    if not temp_id:
        return None
    for fn in os.listdir(UPLOAD_FOLDER):
        if fn.startswith(temp_id):
            return os.path.join(UPLOAD_FOLDER, fn)
    return None

def _parse_number(v):
    """清理字符串并尝试解析为 float，失败返回 None"""
    if v is None:
        return None
    try:
        s = str(v).strip()
        s = re.sub(r'[^\d\.\-]', '', s)
        if s == '':
            return None
        return float(s)
    except Exception:
        return None

# --- 新增：基础变量与工具函数（必须） ---
# 上传/临时相关常量（放在 imports 之后、app.config 使用之前）
MAX_UPLOAD_SIZE = 100 * 1024 * 1024  # 100 MB
ALLOWED_IMPORT_EXT = {'csv', 'xls', 'xlsx'}
TMP_PREFIX = 'zhiguan_'
TMP_EXPIRE_SECONDS = 24 * 3600  # 24 小时

# 轻量异步执行器
_IMPORT_EXECUTOR = concurrent.futures.ThreadPoolExecutor(max_workers=2)
# 导入任务并发控制（避免同时太多大文件）
_IMPORT_LOCK = threading.Lock()

def cleanup_tmp_files():
    tmpdir = tempfile.gettempdir()
    now = time.time()
    for name in os.listdir(tmpdir):
        if name.startswith(TMP_PREFIX) and name.endswith('.pkl'):
            path = os.path.join(tmpdir, name)
            try:
                if now - os.path.getmtime(path) > TMP_EXPIRE_SECONDS:
                    os.remove(path)
                    logger.info(f"已清理旧临时文件: {path}")
            except Exception:
                pass

cleanup_tmp_files()

# 登录装饰器（使用 wraps 保留函数元信息）
def login_required(f):
    @wraps(f)
    def wrap(*args, **kwargs):
        if 'user' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return wrap

# 安全表达式解释器（若安装 asteval，使用之；否则受限 eval）
if _AEVAL_AVAILABLE:
    aeval = Interpreter(usersyms={'sum': sum, 'avg': lambda x: sum(x)/len(x) if x else 0})
else:
    aeval = None
    logger.warning("asteval 未安装，formula 将使用受限 eval（推荐安装 asteval）")

def normalize_product_name(name: str) -> str:
    """
    归一化产品名（保留括号内说明）。
    去除常见单位/词，保留中文字母数字和括号，合并多空格，返回小写字符串。
    """
    if not name:
        return ''
    s = str(name).strip().lower()
    # 去掉括号及其内容
    s = re.sub(r'[\(\（].*?[\)\）]', '', s)
    # 去掉单位/常见词
    s = re.sub(r'\b(kg|g|斤|箱|袋|包|克|千克|公斤|kg\.)\b', '', s)
    # 仅保留字母数字、中文、空格和括号
    s = re.sub(r'[^\w\s\(\)\u4e00-\u9fff]', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s

def fuzzy_ratio(a: str, b: str) -> float:
    if _rf_fuzz is not None:
        try:
            return _rf_fuzz.ratio(a, b)
        except Exception:
            pass
    # 回退到 difflib（返回 0-100）
    return int(difflib.SequenceMatcher(None, a, b).ratio() * 100)

def fuzzy_match_product(normalized_name: str, threshold: int = 90, limit: int = 3):
    """
    在 products 表中按 normalized_name 先做精确查找，再模糊匹配返回候选列表：
    返回 [(product_id, name, normalized_name, score), ...]
    """
    if not normalized_name:
        return []
    conn = _get_conn()
    cur = conn.cursor()
    # 精确匹配 normalized_name
    cur.execute("SELECT id, name, COALESCE(normalized_name, '') FROM products WHERE normalized_name = ?", (normalized_name,))
    row = cur.fetchone()
    if row:
        conn.close()
        return [(row[0], row[1], row[2], 100)]
    # 拉取所有候选 normalized_name
    cur.execute("SELECT id, name, COALESCE(normalized_name, '') FROM products")
    rows = cur.fetchall()
    cand = []
    for r in rows:
        pid, pname, pnorm = r
        score = fuzzy_ratio(normalized_name, pnorm or pname or '')
        if score >= threshold:
            cand.append((pid, pname, pnorm, int(score)))
    cand.sort(key=lambda x: x[3], reverse=True)
    conn.close()
    return cand[:limit]

def to_bid_month(s: str, global_month: str = None) -> str:
    """
    将字符串或全局值转为 YYYY-MM，优先解析行值，失败时使用 global_month。
    """
    if s:
        s = str(s).strip()
        for fmt in ('%Y-%m-%d','%Y/%m/%d','%Y.%m.%d','%Y-%m','%Y/%m','%Y.%m','%Y'):
            try:
                dt = datetime.strptime(s, fmt)
                return dt.strftime('%Y-%m')
            except Exception:
                continue
    if global_month:
        try:
            gm = str(global_month).strip()
            if len(gm) >= 7:
                return gm[:7]
            dt = datetime.strptime(gm, '%Y%m')
            return dt.strftime('%Y-%m')
        except Exception:
            pass
    return ''

# ========== 路由：登录/登出 ==========
@app.route('/')
def index():
    """首页跳转到登录页或仪表盘"""
    if 'user' in session:
        return redirect(url_for('dashboard'))
    else:
        return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username') or 'user'
        session['user'] = username
        flash('已登录：' + username)
        return redirect(url_for('index'))
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.pop('user', None)
    flash('已登出')
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    try:
        conn = get_db()
        cur = conn.cursor()
        
        # 获取各模块统计数据
        stats = {}
        
        # 智能报价统计
        try:
            cur.execute('SELECT COUNT(*) FROM quotes')
            stats['quotes_count'] = cur.fetchone()[0]
        except:
            stats['quotes_count'] = 0
        
        # 客户统计
        try:
            cur.execute('SELECT COUNT(*) FROM customers')
            stats['customers_count'] = cur.fetchone()[0]
        except:
            stats['customers_count'] = 0
        
        # 供应商统计
        try:
            cur.execute('SELECT COUNT(*) FROM suppliers')
            stats['suppliers_count'] = cur.fetchone()[0]
        except:
            stats['suppliers_count'] = 0
        
        # 销售订单统计
        try:
            cur.execute('SELECT COUNT(*) FROM sales_orders')
            stats['sales_orders_count'] = cur.fetchone()[0]
        except:
            stats['sales_orders_count'] = 0
        
        # 采购订单统计
        try:
            cur.execute('SELECT COUNT(*) FROM purchase_orders')
            stats['purchase_orders_count'] = cur.fetchone()[0]
        except:
            stats['purchase_orders_count'] = 0
        
        # 旧订单统计（兼容）
        try:
            cur.execute('SELECT COUNT(*) FROM orders')
            stats['orders_count'] = cur.fetchone()[0]
        except:
            stats['orders_count'] = 0
        
        conn.close()
        
        return render_template('dashboard.html', stats=stats)
    except Exception as e:
        logger.exception("Dashboard错误")
        return f"Dashboard页面: {str(e)}", 500


# ========== 智能报价（页面/数据/批量） ==========
@app.route('/smart_quote')
@login_required
def smart_quote():
    conn = get_db()
    product_filter = request.args.get('product', '')
    date_filter = request.args.get('date', '')
    company_filter = request.args.get('company', '')
    query = 'SELECT * FROM quotes WHERE 1=1'
    params = []
    if product_filter:
        query += ' AND product LIKE ?'
        params.append(f'%{product_filter}%')
    if date_filter:
        query += ' AND bid_date LIKE ?'
        params.append(f'%{date_filter}%')
    if company_filter:
        query += ' AND company LIKE ?'
        params.append(f'%{company_filter}%')
    quotes = conn.execute(query, params).fetchall()
    conn.close()
    return render_template('smart_quote.html', quotes=quotes)

@app.route('/smart_quote/data')
@login_required
def smart_quote_data():
    try:
        conn = get_db()
        rows = conn.execute('SELECT * FROM quotes').fetchall()
        conn.close()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        logger.exception("smart_quote/data 错误")
        return jsonify({"error": str(e)}), 500

# 批量导入（上传 -> 映射 -> 导入）
# 完全替换第240-347行的内容

@app.route('/smart_quote/bulk', methods=['GET','POST'])
@login_required
def smart_quote_bulk():
    if request.method == 'GET':
        return render_template('smart_quote_bulk.html')
    
    # POST请求处理 - 处理文件上传
    try:
        if 'file' not in request.files:
            flash('未选择文件')
            return redirect(request.url)
        
        file = request.files['file']
        if file.filename == '':
            flash('未选择文件')
            return redirect(request.url)
        
        if file and file.filename.lower().endswith(('.csv', '.xlsx', '.xls')):
            # 生成临时文件ID
            temp_id = f"quote_{int(time.time())}_{secrets.token_hex(8)}"
            filename = secure_filename(file.filename)
            filepath = os.path.join(UPLOAD_FOLDER, f"{temp_id}_{filename}")
            file.save(filepath)
            
            # 读取文件预览数据
            try:
                if filename.lower().endswith('.csv'):
                    df = pd.read_csv(filepath, encoding='utf-8', nrows=5)
                else:
                    df = pd.read_excel(filepath, nrows=5)
                
                columns = df.columns.tolist()
                
                # 添加产品名称清洗功能
                def clean_product_name(name):
                    """清洗产品名称：删除-及后面的内容"""
                    if pd.isna(name):
                        return ''
                    name_str = str(name).strip()
                    if '-' in name_str:
                        return name_str.split('-')[0].strip()
                    return name_str
                
                # 预处理数据：清洗所有列的产品名称（因为不知道哪列是产品名称）
                for col in columns:
                    if df[col].dtype == 'object':  # 只处理文本列
                        df[col] = df[col].apply(clean_product_name)
                
                preview_data = df.head().to_dict('records')
                
                # 获取已有公司列表
                conn = get_db()
                companies = conn.execute('SELECT DISTINCT company FROM quotes WHERE company IS NOT NULL AND company != ""').fetchall()
                company_list = [row['company'] for row in companies]
                conn.close()
                
                # 生成年份选项（当前年份前后5年）
                current_year = datetime.now().year
                years = list(range(current_year - 5, current_year + 6))
                months = list(range(1, 13))
                
                return render_template('smart_quote_bulk.html', 
                                     show_mapping=True,
                                     temp_id=temp_id,
                                     columns=columns,
                                     preview_data=preview_data,
                                     companies=company_list,
                                     years=years,
                                     months=months,
                                     current_year=current_year,
                                     current_month=datetime.now().month)
                
            except Exception as e:
                flash(f'文件读取失败: {str(e)}')
                return redirect(request.url)
        else:
            flash('仅支持 CSV、Excel 文件')
            return redirect(request.url)
            
    except Exception as e:
        flash(f'上传失败: {str(e)}')
        return redirect(request.url)

@app.route('/smart_quote/import', methods=['POST'])
@login_required
def smart_quote_import():
    """处理智能报价导入"""
    try:
        temp_id = request.form.get('temp_id')
        product_col = request.form.get('product_col')
        price_col = request.form.get('price_col')
        qty_col = request.form.get('qty_col')
        company_name = request.form.get('company_name')
        bid_year = request.form.get('bid_year')
        bid_month = request.form.get('bid_month')
        
        if not all([temp_id, product_col, price_col, company_name, bid_year, bid_month]):
            flash('请填写所有必填字段')
            return redirect(url_for('smart_quote_bulk'))
        
        # 构建日期
        bid_date = f"{bid_year}-{int(bid_month):02d}-01"
        
        # 查找临时文件
        filepath = _locate_temp_file(temp_id)
        if not filepath:
            flash('临时文件不存在或已过期')
            return redirect(url_for('smart_quote_bulk'))
        
        # 执行导入
        result = process_smart_quote_import_new(
            filepath, product_col, price_col, qty_col,
            company_name, bid_date, 'overwrite'
        )
        
        if result['success']:
            flash(f"导入成功！成功：{result['success_count']}，跳过：{result['skip_count']}，失败：{result['error_count']}")
        else:
            flash(f"导入失败：{result['error']}")
        
        return redirect(url_for('smart_quote'))
        
    except Exception as e:
        logger.exception("导入处理失败")
        flash(f'导入处理失败: {str(e)}')
        return redirect(url_for('smart_quote_bulk'))
        
def process_smart_quote_import_new(filepath, product_col, price_col, qty_col, 
                                  company_name, bid_date, conflict_mode):
    """新的导入处理逻辑 - 统一公司和日期"""
    
    # 产品名称清洗函数
    def clean_product_name(name):
        """清洗产品名称：删除-及后面的内容"""
        if pd.isna(name):
            return ''
        name_str = str(name).strip()
        if '-' in name_str:
            return name_str.split('-')[0].strip()
        return name_str
    
    try:
        # 读取文件
        if filepath.lower().endswith('.csv'):
            df = pd.read_csv(filepath, encoding='utf-8')
        else:
            df = pd.read_excel(filepath)
        
        conn = get_db()
        cur = conn.cursor()
        
        success_count = 0
        skip_count = 0
        error_count = 0
        errors = []
        
        for idx, row in df.iterrows():
            try:
                # 从Excel读取的数据并清洗产品名称
                raw_product_name = str(row.get(product_col, '')).strip()
                product_name = clean_product_name(raw_product_name)
                
                price_value = _parse_number(row.get(price_col))
                qty_value = int(_parse_number(row.get(qty_col)) or 1) if qty_col else 1
                
                # 必填字段验证
                if not product_name or price_value is None:
                    error_count += 1
                    errors.append(f'第{idx+2}行: 产品名称或中标价格为空')
                    continue
                
                # 价格合理性验证
                if price_value <= 0:
                    error_count += 1
                    errors.append(f'第{idx+2}行: 中标价格必须大于0')
                    continue
                
                # 修复：完整的冲突检查逻辑
                existing = cur.execute(
                    'SELECT id, price FROM quotes WHERE product=? AND company=? AND bid_date=?',
                    (product_name, company_name, bid_date)
                ).fetchone()
                
                if existing:
                    if conflict_mode == 'skip':
                        skip_count += 1
                        continue  # 跳过这条记录
                    elif conflict_mode == 'overwrite' or conflict_mode == 'replace':
                        # 更新现有记录
                        cur.execute('''
                            UPDATE quotes SET price=?, qty=?, remarks=? WHERE id=?
                        ''', (price_value, qty_value, f'更新_{datetime.now().strftime("%Y%m%d")}', existing[0]))
                        success_count += 1
                        continue
                    else:
                        # 默认跳过
                        skip_count += 1
                        continue
                
                # 插入新记录（只有在没有冲突或已处理冲突时才执行）
                cur.execute('''
                    INSERT INTO quotes (product, company, price, qty, bid_date, remarks)
                    VALUES (?, ?, ?, ?, ?, ?)
                ''', (product_name, company_name, price_value, qty_value, bid_date, 
                     f'批量导入_{datetime.now().strftime("%Y%m%d")}'))
                
                success_count += 1
                
            except Exception as e:
                error_count += 1
                errors.append(f'第{idx+2}行: {str(e)}')
                logger.exception(f"导入第{idx+1}行时出错")
        
        conn.commit()
        conn.close()
        
        # 清理临时文件
        try:
            os.remove(filepath)
        except:
            pass
        
        return {
            'success': True,
            'success_count': success_count,
            'skip_count': skip_count,
            'error_count': error_count,
            'errors': errors[:10]
        }
        
    except Exception as e:
        return {'success': False, 'error': str(e)}

@app.route('/api/companies')
@login_required
def api_companies():
    q = request.args.get('q', '')
    conn = get_db()
    if q:
        rows = conn.execute('SELECT name FROM customers WHERE name LIKE ?', (f'%{q}%',)).fetchall()
    else:
        rows = conn.execute('SELECT name FROM customers').fetchall()
    conn.close()
    return jsonify([r['name'] for r in rows])

@app.route('/api/smart_quotes/<int:qid>')
@login_required
def api_smart_quote_get(qid):
    conn = get_db()
    row = conn.execute('SELECT * FROM quotes WHERE id=?', (qid,)).fetchone()
    conn.close()
    if not row:
        return jsonify({"error": "not found"}), 404
    return jsonify(dict(row))

@app.route('/api/smart_quotes/edit', methods=['POST'])
@login_required
def api_smart_quote_edit():
    payload = request.get_json() or {}
    qid = payload.get('id')
    data = payload.get('data', {})
    if not qid:
        return jsonify({"success": False, "error": "missing id"}), 400
    conn = get_db()
    conn.execute('UPDATE quotes SET company=?, price=?, qty=?, bid_date=?, remarks=?, default_bid=? WHERE id=?',
                 (data.get('company'), data.get('price'), data.get('qty'), data.get('bid_date'), data.get('remarks'), data.get('default_bid'), qid))
    conn.commit()
    conn.close()
    return jsonify({"success": True})

@app.route('/api/smart_quotes/delete', methods=['POST'])
@login_required
def api_smart_quote_delete():
    payload = request.get_json() or {}
    qid = payload.get('id')
    if not qid:
        return jsonify({"success": False, "error": "missing id"}), 400
    conn = get_db()
    conn.execute('DELETE FROM quotes WHERE id=?', (qid,))
    conn.commit()
    conn.close()
    return jsonify({"success": True})

# 找到 api_smart_quote_search 函数，大约在第350行左右

@app.route('/api/smart_quote/search', methods=['POST'])
@login_required
def api_smart_quote_search():
    """智能报价搜索API - 支持分页和去重"""
    try:
        data = request.get_json()
        if not data:
            data = {}
        
        # 搜索条件
        product = data.get('product', '').strip()
        company = data.get('company', '').strip()  
        date_start = data.get('date_start', '').strip()
        date_end = data.get('date_end', '').strip()
        price_min = data.get('price_min')
        price_max = data.get('price_max')
        
        # 分页参数
        page = int(data.get('page', 1))
        page_size = int(data.get('page_size', 20))
        offset = (page - 1) * page_size
        
        conn = get_db()
        cur = conn.cursor()
        
        # 构建查询条件
        where_conditions = ['1=1']
        params = []
        
        if product:
            where_conditions.append('product LIKE ?')
            params.append(f'%{product}%')
        
        if company:
            where_conditions.append('company = ?')
            params.append(company)
            
        if date_start:
            where_conditions.append('bid_date >= ?')
            params.append(f'{date_start}-01')
            
        if date_end:
            try:
                year, month = date_end.split('-')
                year, month = int(year), int(month)
                
                if month == 12:
                    next_year, next_month = year + 1, 1
                else:
                    next_year, next_month = year, month + 1
                    
                where_conditions.append('bid_date < ?')
                params.append(f'{next_year:04d}-{next_month:02d}-01')
            except (ValueError, TypeError) as e:
                logger.warning(f"时间格式错误: {date_end}, 错误: {e}")
            
        if price_min is not None:
            where_conditions.append('price >= ?')
            params.append(float(price_min))
            
        if price_max is not None:
            where_conditions.append('price <= ?')
            params.append(float(price_max))
        
        where_clause = ' AND '.join(where_conditions)
        
        # 修复：使用CTE和ROW_NUMBER进行去重，保留最新记录
        count_sql = f'''
            WITH deduplicated AS (
                SELECT id, product, company, price, qty, bid_date, remarks,
                       ROW_NUMBER() OVER (PARTITION BY product, company, bid_date ORDER BY id DESC) as rn
                FROM quotes 
                WHERE {where_clause}
            )
            SELECT COUNT(*) FROM deduplicated WHERE rn = 1
        '''
        total = cur.execute(count_sql, params).fetchone()[0]
        
        # 查询去重后的分页数据
        data_sql = f'''
            WITH deduplicated AS (
                SELECT id, product, company, price, qty, bid_date, remarks,
                       ROW_NUMBER() OVER (PARTITION BY product, company, bid_date ORDER BY id DESC) as rn
                FROM quotes 
                WHERE {where_clause}
            )
            SELECT id, product, company, price, qty, bid_date, remarks
            FROM deduplicated 
            WHERE rn = 1
            ORDER BY bid_date DESC, id DESC 
            LIMIT ? OFFSET ?
        '''
        results = cur.execute(data_sql, params + [page_size, offset]).fetchall()
        
        # 获取公司列表
        companies_sql = 'SELECT DISTINCT company FROM quotes WHERE company IS NOT NULL AND company != "" ORDER BY company'
        companies = cur.execute(companies_sql).fetchall()
        
        conn.close()
        
        # 转换结果为字典格式
        data_list = []
        for row in results:
            data_list.append({
                'id': row[0],
                'product': row[1],
                'company': row[2],
                'price': row[3],
                'qty': row[4],
                'bid_date': row[5],
                'remarks': row[6] if len(row) > 6 else None
            })
        
        result = {
            'success': True,
            'data': data_list,
            'pagination': {
                'current_page': page,
                'page_size': page_size,
                'total_records': total,
                'total_pages': (total + page_size - 1) // page_size
            },
            'options': {
                'companies': [row[0] for row in companies]
            }
        }
        
        return jsonify(result)
        
    except Exception as e:
        logger.exception("搜索错误")
        return jsonify({'success': False, 'error': str(e)}), 500

# ========== 公式（安全执行） ==========
@app.route('/formula', methods=['GET', 'POST'])
@login_required
def formula():
    if request.method == 'POST':
        formula_text = request.form.get('formula', '')
        # 限制字符集（简单防护）
        allowed = set("0123456789+-*/()., _[]abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ")
        if not set(formula_text) <= allowed:
            flash('公式包含不允许的字符')
        else:
            try:
                if aeval is not None:
                    result = aeval(formula_text)
                else:
                    # 受限 eval（仅提供 sum/avg）
                    result = eval(formula_text, {"__builtins__": {}}, {"sum": sum, "avg": lambda x: sum(x)/len(x) if x else 0})
                flash(f'预览结果: {result}')
            except Exception as e:
                logger.exception("公式解析失败")
                flash('公式解析失败')
        conn = get_db()
        conn.execute('INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)', ('formula', formula_text))
        conn.commit()
        conn.close()
    return render_template('formula.html')

@app.route('/calculation')
@login_required
def calculation_analysis():
    # 获取所有公司名称（用于下拉选择）
    conn = get_db()
    companies = conn.execute('SELECT DISTINCT company FROM quotes WHERE company IS NOT NULL AND company != ""').fetchall()
    conn.close()
    
    return render_template('calculation_analysis.html', companies=companies)

# ========== 订单模块（含 JSON API） ==========

@app.route('/order/data')
@login_required
def order_data():
    try:
        conn = get_db()
        data = conn.execute('SELECT * FROM orders').fetchall()
        conn.close()
        return jsonify([dict(r) for r in data])
    except Exception as e:
        logger.exception("order/data 错误")
        return jsonify({"error": str(e)}), 500

@app.route('/order/add', methods=['POST'])
@login_required
def add_order():
    try:
        if request.is_json:
            data = request.get_json()
        else:
            data = request.form.to_dict()
        details_json = data.get('details_json') or '[]'
        details = json.loads(details_json)
        total_price = sum(float(d.get('price',0))*float(d.get('qty',0)) for d in details)
        conn = get_db()
        conn.execute('INSERT INTO orders (type, customer, date, total_price, status, details_count) VALUES (?, ?, ?, ?, ?, ?)',
                     (data.get('type'), data.get('customer'), data.get('date'), total_price, data.get('status'), len(details)))
        conn.commit()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        logger.exception("新增订单失败")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/order/update/<int:oid>', methods=['POST'])
@login_required
def update_order(oid):
    try:
        if request.is_json:
            data = request.get_json()
        else:
            data = request.form.to_dict()
        details = json.loads(data.get('details_json','[]'))
        total_price = sum(float(d.get('price',0))*float(d.get('qty',0)) for d in details)
        conn = get_db()
        conn.execute('UPDATE orders SET type=?, customer=?, date=?, total_price=?, status=?, details_count=? WHERE id=?',
                     (data.get('type'), data.get('customer'), data.get('date'), total_price, data.get('status'), len(details), oid))
        conn.commit()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        logger.exception("更新订单失败")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/order/delete/<int:oid>')
@login_required
def delete_order(oid):
    conn = get_db()
    conn.execute('DELETE FROM order_details WHERE order_id=?', (oid,))
    conn.execute('DELETE FROM orders WHERE id=?', (oid,))
    conn.commit()
    conn.close()
    flash('订单删除成功')
    return redirect(url_for('order'))

# ========== 整体导航/错误处理（新增，确保页面不丢） ==========
@app.errorhandler(404)
def page_not_found(e):
    # 简单统一处理：闪现消息并返回仪表盘
    flash('请求的页面未找到，已返回仪表盘')
    return redirect(url_for('dashboard'))

@app.route('/pivot')
@login_required
def pivot():
    # 若需从 DB 读取数据，可替换下面的空数组
    try:
        conn = get_db()
        rows = conn.execute('SELECT * FROM quotes LIMIT 100').fetchall()
        conn.close()
        data = [dict(r) for r in rows]
    except Exception:
        data = []
    return render_template('pivot.html', data=data)

@app.route('/settings')
@login_required
def settings():
    try:
        conn = get_db()
        cur = conn.cursor()
        
        # 获取所有设置
        cur.execute('SELECT * FROM settings')
        settings_dict = {row['key']: row['value'] for row in cur.fetchall()}
        
        conn.close()
        return render_template('settings.html', settings=settings_dict)
    except Exception as e:
        return f"设置页面: {str(e)}", 500
    
    # 添加缺失的import_upload_page路由
@app.route('/import')
@login_required
def import_upload_page():
    return render_template('import_upload.html')

# 启动时确保表存在
def ensure_tables():
    conn = get_db()
    cur = conn.cursor()
    cur.execute('''CREATE TABLE IF NOT EXISTS quotes 
                    (id INTEGER PRIMARY KEY, product TEXT, company TEXT, price REAL, qty INTEGER, 
                     bid_date TEXT, remarks TEXT, default_bid REAL)''')
    cur.execute('''CREATE TABLE IF NOT EXISTS orders 
                    (id INTEGER PRIMARY KEY, type TEXT, customer TEXT, date TEXT, total_price REAL, 
                     status TEXT, details_count INTEGER)''')
    cur.execute('''CREATE TABLE IF NOT EXISTS settings (key TEXT PRIMARY KEY, value TEXT)''')
    cur.execute('''CREATE TABLE IF NOT EXISTS customers 
                    (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT UNIQUE)''')
    # 创建 products 表并保留 normalized_name 列（兼容旧表）
    cur.execute('''CREATE TABLE IF NOT EXISTS products 
                    (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT UNIQUE, normalized_name TEXT, created_at TEXT)''')
    cur.execute('''CREATE TABLE IF NOT EXISTS order_details 
                    (id INTEGER PRIMARY KEY AUTOINCREMENT, order_id INTEGER, product_id INTEGER, qty INTEGER, price REAL)''')
    cur.execute('''CREATE TABLE IF NOT EXISTS warehouses (id INTEGER PRIMARY KEY, name TEXT)''')
    cur.execute('''CREATE TABLE IF NOT EXISTS categories 
                    (id INTEGER PRIMARY KEY, name TEXT)''')
    cur.execute('''CREATE TABLE IF NOT EXISTS inventory 
                    (id INTEGER PRIMARY KEY, product_id INTEGER, warehouse_id INTEGER, category_id INTEGER, 
                     qty INTEGER, last_update TEXT)''')

    # price_meta 与导入任务/错误表（代码中使用到，确保存在）
    cur.execute('''CREATE TABLE IF NOT EXISTS price_meta
                    (id INTEGER PRIMARY KEY AUTOINCREMENT, product_id INTEGER, bid_month TEXT, company TEXT, price REAL, price_type TEXT, created_at TEXT)''')
    cur.execute('''CREATE TABLE IF NOT EXISTS import_tasks
                    (id TEXT PRIMARY KEY, temp_id TEXT, filename TEXT, mapping TEXT, conflict_mode TEXT, status TEXT, created_at TEXT, updated_at TEXT, total INTEGER, success INTEGER, failed INTEGER)''')
    cur.execute('''CREATE TABLE IF NOT EXISTS import_errors
                    (id INTEGER PRIMARY KEY AUTOINCREMENT, task_id TEXT, row_no INTEGER, raw TEXT, error_msg TEXT)''')

    # ========== 订单系统数据库表 ==========
    
    # 1. 客户表
    cur.execute('''CREATE TABLE IF NOT EXISTS customers (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        customer_code VARCHAR(20) UNIQUE NOT NULL,
        customer_name VARCHAR(100) NOT NULL,
        contact_person VARCHAR(50),
        contact_phone VARCHAR(20),
        address TEXT,
        remarks TEXT,
        create_time DATETIME DEFAULT CURRENT_TIMESTAMP,
        update_time DATETIME DEFAULT CURRENT_TIMESTAMP
    )''')
    
    # 2. 供应商表
    cur.execute('''CREATE TABLE IF NOT EXISTS suppliers (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        supplier_code VARCHAR(20) UNIQUE NOT NULL,
        supplier_name VARCHAR(100) NOT NULL,
        contact_person VARCHAR(50),
        contact_phone VARCHAR(20),
        address TEXT,
        remarks TEXT,
        create_time DATETIME DEFAULT CURRENT_TIMESTAMP,
        update_time DATETIME DEFAULT CURRENT_TIMESTAMP
    )''')
    
    # 3. 销售订单表
    cur.execute('''CREATE TABLE IF NOT EXISTS sales_orders (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        order_code VARCHAR(20) UNIQUE NOT NULL,
        customer_id INTEGER NOT NULL,
        customer_name VARCHAR(100) NOT NULL,
        order_date DATE NOT NULL,
        delivery_date DATE,
        total_amount DECIMAL(10,2) DEFAULT 0,
        status VARCHAR(20) DEFAULT '待确认',
        remarks TEXT,
        create_time DATETIME DEFAULT CURRENT_TIMESTAMP,
        update_time DATETIME DEFAULT CURRENT_TIMESTAMP,
        create_user VARCHAR(50),
        FOREIGN KEY (customer_id) REFERENCES customers(id)
    )''')
    
    # 4. 销售订单明细表
    cur.execute('''CREATE TABLE IF NOT EXISTS sales_order_items (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        order_id INTEGER NOT NULL,
        smart_quote_id INTEGER,
        product_name VARCHAR(100) NOT NULL,
        category VARCHAR(50),
        specification VARCHAR(50),
        unit VARCHAR(20) NOT NULL,
        quantity DECIMAL(10,2) NOT NULL,
        price DECIMAL(10,2) NOT NULL,
        amount DECIMAL(10,2) NOT NULL,
        remarks TEXT,
        create_time DATETIME DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (order_id) REFERENCES sales_orders(id) ON DELETE CASCADE,
        FOREIGN KEY (smart_quote_id) REFERENCES quotes(id) ON DELETE SET NULL
    )''')
    
    # 5. 采购订单表
    cur.execute('''CREATE TABLE IF NOT EXISTS purchase_orders (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        order_code VARCHAR(20) UNIQUE NOT NULL,
        supplier_id INTEGER NOT NULL,
        supplier_name VARCHAR(100) NOT NULL,
        order_date DATE NOT NULL,
        expected_date DATE,
        total_amount DECIMAL(10,2) DEFAULT 0,
        status VARCHAR(20) DEFAULT '待确认',
        remarks TEXT,
        create_time DATETIME DEFAULT CURRENT_TIMESTAMP,
        update_time DATETIME DEFAULT CURRENT_TIMESTAMP,
        create_user VARCHAR(50),
        FOREIGN KEY (supplier_id) REFERENCES suppliers(id)
    )''')
    
    # 6. 采购订单明细表
    cur.execute('''CREATE TABLE IF NOT EXISTS purchase_order_items (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        order_id INTEGER NOT NULL,
        product_name VARCHAR(100) NOT NULL,
        category VARCHAR(50),
        specification VARCHAR(50),
        unit VARCHAR(20) NOT NULL,
        quantity DECIMAL(10,2) NOT NULL,
        price DECIMAL(10,2) NOT NULL,
        amount DECIMAL(10,2) NOT NULL,
        remarks TEXT,
        create_time DATETIME DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (order_id) REFERENCES purchase_orders(id) ON DELETE CASCADE
    )''')
    
    # 7. 分拣标签表
    cur.execute('''CREATE TABLE IF NOT EXISTS picking_labels (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        label_code VARCHAR(20) UNIQUE NOT NULL,
        order_id INTEGER NOT NULL,
        order_code VARCHAR(20),
        customer_name VARCHAR(100),
        product_name VARCHAR(100) NOT NULL,
        category VARCHAR(50),
        specification VARCHAR(50),
        unit VARCHAR(20),
        quantity DECIMAL(10,2) NOT NULL,
        remarks TEXT,
        status VARCHAR(20) DEFAULT '待分拣',
        picker VARCHAR(50),
        pick_time DATETIME,
        create_time DATETIME DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (order_id) REFERENCES sales_orders(id) ON DELETE CASCADE
    )''')

        # ========== 新增：库存管理表 ==========
    
    # 库存主表
    cur.execute('''CREATE TABLE IF NOT EXISTS inventory (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        product_name VARCHAR(100) NOT NULL,
        category VARCHAR(50),
        specification VARCHAR(100),
        unit VARCHAR(20) DEFAULT '件',
        current_stock DECIMAL(10,2) DEFAULT 0,
        safe_stock DECIMAL(10,2) DEFAULT 0,
        warehouse_location VARCHAR(50),
        remarks TEXT,
        create_time DATETIME DEFAULT CURRENT_TIMESTAMP,
        update_time DATETIME DEFAULT CURRENT_TIMESTAMP
    )''')
    
    # 入库记录表
    cur.execute('''CREATE TABLE IF NOT EXISTS inbound_records (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        record_code VARCHAR(20) UNIQUE NOT NULL,
        inbound_type VARCHAR(20) DEFAULT '采购入库',
        product_name VARCHAR(100) NOT NULL,
        category VARCHAR(50),
        specification VARCHAR(100),
        quantity DECIMAL(10,2) NOT NULL,
        unit VARCHAR(20),
        purchase_order_code VARCHAR(20),
        supplier_name VARCHAR(100),
        inbound_date DATE NOT NULL,
        operator VARCHAR(50),
        remarks TEXT,
        create_time DATETIME DEFAULT CURRENT_TIMESTAMP
    )''')
    
    # 出库记录表
    cur.execute('''CREATE TABLE IF NOT EXISTS outbound_records (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        record_code VARCHAR(20) UNIQUE NOT NULL,
        outbound_type VARCHAR(20) DEFAULT '销售出库',
        product_name VARCHAR(100) NOT NULL,
        category VARCHAR(50),
        specification VARCHAR(100),
        quantity DECIMAL(10,2) NOT NULL,
        unit VARCHAR(20),
        sales_order_code VARCHAR(20),
        customer_name VARCHAR(100),
        outbound_date DATE NOT NULL,
        operator VARCHAR(50),
        remarks TEXT,
        create_time DATETIME DEFAULT CURRENT_TIMESTAMP
    )''')
    
    # 分拣标签表（修复字段）
    cur.execute('''CREATE TABLE IF NOT EXISTS picking_labels (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        label_code VARCHAR(20) UNIQUE NOT NULL,
        order_id INTEGER NOT NULL,
        order_code VARCHAR(20),
        customer_name VARCHAR(100),
        product_name VARCHAR(100) NOT NULL,
        category VARCHAR(50),
        specification VARCHAR(50),
        quantity DECIMAL(10,2) NOT NULL,
        unit VARCHAR(20) DEFAULT '件',
        delivery_date DATE,
        label_status VARCHAR(20) DEFAULT '待打印',
        print_count INTEGER DEFAULT 0,
        remarks TEXT,
        create_user VARCHAR(50),
        create_time DATETIME DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (order_id) REFERENCES sales_orders(id) ON DELETE CASCADE
    )''')
    
    conn.commit()
    
    # 创建索引
    try:
        cur.execute('CREATE INDEX IF NOT EXISTS idx_customers_name ON customers(customer_name)')
        cur.execute('CREATE INDEX IF NOT EXISTS idx_customers_code ON customers(customer_code)')
        cur.execute('CREATE INDEX IF NOT EXISTS idx_suppliers_name ON suppliers(supplier_name)')
        cur.execute('CREATE INDEX IF NOT EXISTS idx_suppliers_code ON suppliers(supplier_code)')
        cur.execute('CREATE INDEX IF NOT EXISTS idx_sales_orders_code ON sales_orders(order_code)')
        cur.execute('CREATE INDEX IF NOT EXISTS idx_sales_orders_date ON sales_orders(order_date)')
        cur.execute('CREATE INDEX IF NOT EXISTS idx_sales_orders_status ON sales_orders(status)')
        cur.execute('CREATE INDEX IF NOT EXISTS idx_sales_orders_customer ON sales_orders(customer_id)')
        cur.execute('CREATE INDEX IF NOT EXISTS idx_purchase_orders_code ON purchase_orders(order_code)')
        cur.execute('CREATE INDEX IF NOT EXISTS idx_purchase_orders_date ON purchase_orders(order_date)')
        cur.execute('CREATE INDEX IF NOT EXISTS idx_purchase_orders_status ON purchase_orders(status)')
        cur.execute('CREATE INDEX IF NOT EXISTS idx_picking_labels_order ON picking_labels(order_id)')
        cur.execute('CREATE INDEX IF NOT EXISTS idx_picking_labels_status ON picking_labels(status)')
    except Exception as e:
        logger.warning(f"创建索引时出现警告: {str(e)}")

    conn.commit()
    # 兼容：若旧的 products 表缺少 normalized_name 列，尝试添加（SQLite 在重复添加时会抛错，捕获忽略）
    try:
        cur.execute("ALTER TABLE products ADD COLUMN normalized_name TEXT")
        cur.execute("ALTER TABLE products ADD COLUMN created_at TEXT")
        conn.commit()
    except Exception:
        pass

    conn.close()

# 确保在模块加载时初始化表（方便直接用 python app.py 启动）
ensure_tables()

# 在 ensure_tables() 函数之后添加（约第 1245 行）

def generate_customer_code():
    """生成客户编号 KH20250127001"""
    conn = get_db()
    cur = conn.cursor()
    try:
        today = datetime.now().strftime('%Y%m%d')
        prefix = f"KH{today}"
        
        cur.execute('''
            SELECT customer_code FROM customers 
            WHERE customer_code LIKE ? 
            ORDER BY customer_code DESC LIMIT 1
        ''', (f'{prefix}%',))
        
        result = cur.fetchone()
        if result:
            last_code = result[0]
            last_num = int(last_code[-3:])
            new_num = last_num + 1
        else:
            new_num = 1
        
        return f"{prefix}{new_num:03d}"
    finally:
        conn.close()


def generate_supplier_code():
    """生成供应商编号 GYS20250127001"""
    conn = get_db()
    cur = conn.cursor()
    try:
        today = datetime.now().strftime('%Y%m%d')
        prefix = f"GYS{today}"
        
        cur.execute('''
            SELECT supplier_code FROM suppliers 
            WHERE supplier_code LIKE ? 
            ORDER BY supplier_code DESC LIMIT 1
        ''', (f'{prefix}%',))
        
        result = cur.fetchone()
        if result:
            last_code = result[0]
            last_num = int(last_code[-3:])
            new_num = last_num + 1
        else:
            new_num = 1
        
        return f"{prefix}{new_num:03d}"
    finally:
        conn.close()


def generate_sales_order_code():
    """生成销售订单编号 XS20250127001"""
    conn = get_db()
    cur = conn.cursor()
    try:
        today = datetime.now().strftime('%Y%m%d')
        prefix = f"XS{today}"
        
        cur.execute('''
            SELECT order_code FROM sales_orders 
            WHERE order_code LIKE ? 
            ORDER BY order_code DESC LIMIT 1
        ''', (f'{prefix}%',))
        
        result = cur.fetchone()
        if result:
            last_code = result[0]
            last_num = int(last_code[-3:])
            new_num = last_num + 1
        else:
            new_num = 1
        
        return f"{prefix}{new_num:03d}"
    finally:
        conn.close()


def generate_purchase_order_code():
    """生成采购订单编号 CG20250127001"""
    conn = get_db()
    cur = conn.cursor()
    try:
        today = datetime.now().strftime('%Y%m%d')
        prefix = f"CG{today}"
        
        cur.execute('''
            SELECT order_code FROM purchase_orders 
            WHERE order_code LIKE ? 
            ORDER BY order_code DESC LIMIT 1
        ''', (f'{prefix}%',))
        
        result = cur.fetchone()
        if result:
            last_code = result[0]
            last_num = int(last_code[-3:])
            new_num = last_num + 1
        else:
            new_num = 1
        
        return f"{prefix}{new_num:03d}"
    finally:
        conn.close()


def generate_picking_label_code():
    """生成分拣标签编号 FJ20250127-001"""
    conn = get_db()
    cur = conn.cursor()
    try:
        today = datetime.now().strftime('%Y%m%d')
        prefix = f"FJ{today}"
        
        cur.execute('''
            SELECT label_code FROM picking_labels 
            WHERE label_code LIKE ? 
            ORDER BY label_code DESC LIMIT 1
        ''', (f'{prefix}%',))
        
        result = cur.fetchone()
        if result:
            last_code = result[0]
            last_num = int(last_code.split('-')[-1])
            new_num = last_num + 1
        else:
            new_num = 1
        
        return f"{prefix}-{new_num:03d}"
    finally:
        conn.close()

# --- API 路由 ---

# 在客户管理 API 之后添加（约第 1480 行之后）

# ========== 供应商管理 API ==========

@app.route('/suppliers')
@login_required
def suppliers_page():
    """供应商管理页面"""
    return render_template('suppliers.html')


@app.route('/api/suppliers', methods=['GET'])
@login_required
def get_suppliers():
    """获取供应商列表（分页、搜索）"""
    try:
        page = int(request.args.get('page', 1))
        page_size = int(request.args.get('page_size', 15))
        search_name = request.args.get('name', '').strip()
        search_phone = request.args.get('phone', '').strip()
        
        conn = get_db()
        cur = conn.cursor()
        
        # 构建查询条件
        where_conditions = []
        params = []
        
        if search_name:
            where_conditions.append("supplier_name LIKE ?")
            params.append(f'%{search_name}%')
        
        if search_phone:
            where_conditions.append("contact_phone LIKE ?")
            params.append(f'%{search_phone}%')
        
        where_clause = ' AND '.join(where_conditions) if where_conditions else '1=1'
        
        # 查询总数
        cur.execute(f"SELECT COUNT(*) FROM suppliers WHERE {where_clause}", params)
        total = cur.fetchone()[0]
        
        # 查询数据
        offset = (page - 1) * page_size
        cur.execute(f'''
            SELECT id, supplier_code, supplier_name, contact_person, 
                   contact_phone, address, remarks, create_time, update_time
            FROM suppliers 
            WHERE {where_clause}
            ORDER BY create_time DESC
            LIMIT ? OFFSET ?
        ''', params + [page_size, offset])
        
        columns = [desc[0] for desc in cur.description]
        items = [dict(zip(columns, row)) for row in cur.fetchall()]
        
        conn.close()
        
        return jsonify({
        'success': True,
        'data': items,
        'total': total,
        'pagination': {
            'page': page,
            'page_size': page_size,
            'total_pages': (total + page_size - 1) // page_size
        }
    })
        
    except Exception as e:
        logger.error(f"获取供应商列表失败: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/suppliers/generate_code', methods=['GET'])
@login_required
def generate_supplier_code_api():
    """生成新的供应商编号"""
    try:
        code = generate_supplier_code()
        return jsonify({'success': True, 'code': code})
    except Exception as e:
        logger.error(f"生成供应商编号失败: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/suppliers/<int:supplier_id>', methods=['GET'])
@login_required
def get_supplier(supplier_id):
    """获取单个供应商详情"""
    try:
        conn = get_db()
        cur = conn.cursor()
        
        cur.execute('''
            SELECT id, supplier_code, supplier_name, contact_person,
                   contact_phone, address, remarks, create_time, update_time
            FROM suppliers WHERE id = ?
        ''', (supplier_id,))
        
        row = cur.fetchone()
        conn.close()
        
        if not row:
            return jsonify({'success': False, 'message': '供应商不存在'}), 404
        
        columns = [desc[0] for desc in cur.description]
        supplier = dict(zip(columns, row))
        
        return jsonify({'success': True, 'data': supplier})
        
    except Exception as e:
        logger.error(f"获取供应商详情失败: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/suppliers', methods=['POST'])
@login_required
def add_supplier():
    """添加供应商"""
    try:
        data = request.get_json()
        
        # 验证必填字段
        if not data.get('supplier_name'):
            return jsonify({'success': False, 'message': '供应商名称不能为空'}), 400
        
        if not data.get('supplier_code'):
            data['supplier_code'] = generate_supplier_code()
        
        conn = get_db()
        cur = conn.cursor()
        
        # 检查编号是否重复
        cur.execute('SELECT id FROM suppliers WHERE supplier_code = ?', (data['supplier_code'],))
        if cur.fetchone():
            conn.close()
            return jsonify({'success': False, 'message': '供应商编号已存在'}), 400
        
        # 插入数据
        cur.execute('''
            INSERT INTO suppliers (supplier_code, supplier_name, contact_person,
                                 contact_phone, address, remarks)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (
            data['supplier_code'],
            data['supplier_name'],
            data.get('contact_person'),
            data.get('contact_phone'),
            data.get('address'),
            data.get('remarks')
        ))
        
        conn.commit()
        supplier_id = cur.lastrowid
        conn.close()
        
        return jsonify({'success': True, 'message': '添加成功', 'id': supplier_id})
        
    except Exception as e:
        logger.error(f"添加供应商失败: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/suppliers/<int:supplier_id>', methods=['PUT'])
@login_required
def update_supplier(supplier_id):
    """更新供应商"""
    try:
        data = request.get_json()
        
        if not data.get('supplier_name'):
            return jsonify({'success': False, 'message': '供应商名称不能为空'}), 400
        
        conn = get_db()
        cur = conn.cursor()
        
        # 检查供应商是否存在
        cur.execute('SELECT id FROM suppliers WHERE id = ?', (supplier_id,))
        if not cur.fetchone():
            conn.close()
            return jsonify({'success': False, 'message': '供应商不存在'}), 404
        
        # 更新数据
        cur.execute('''
            UPDATE suppliers 
            SET supplier_name = ?, contact_person = ?, contact_phone = ?,
                address = ?, remarks = ?, update_time = CURRENT_TIMESTAMP
            WHERE id = ?
        ''', (
            data['supplier_name'],
            data.get('contact_person'),
            data.get('contact_phone'),
            data.get('address'),
            data.get('remarks'),
            supplier_id
        ))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': '更新成功'})
        
    except Exception as e:
        logger.error(f"更新供应商失败: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/suppliers/<int:supplier_id>', methods=['DELETE'])
@login_required
def delete_supplier(supplier_id):
    """删除供应商"""
    try:
        conn = get_db()
        cur = conn.cursor()
        
        # 检查是否有关联订单
        cur.execute('SELECT COUNT(*) FROM purchase_orders WHERE supplier_id = ?', (supplier_id,))
        order_count = cur.fetchone()[0]
        
        if order_count > 0:
            conn.close()
            return jsonify({
                'success': False, 
                'message': f'该供应商有 {order_count} 个关联采购订单，无法删除'
            }), 400
        
        # 删除供应商
        cur.execute('DELETE FROM suppliers WHERE id = ?', (supplier_id,))
        
        if cur.rowcount == 0:
            conn.close()
            return jsonify({'success': False, 'message': '供应商不存在'}), 404
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': '删除成功'})
        
    except Exception as e:
        logger.error(f"删除供应商失败: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/suppliers/export', methods=['GET'])
@login_required
def export_suppliers():
    """导出供应商数据到Excel"""
    try:
        import io
        from datetime import datetime
        
        search_name = request.args.get('name', '').strip()
        search_phone = request.args.get('phone', '').strip()
        
        conn = get_db()
        cur = conn.cursor()
        
        # 构建查询条件
        where_conditions = []
        params = []
        
        if search_name:
            where_conditions.append("supplier_name LIKE ?")
            params.append(f'%{search_name}%')
        
        if search_phone:
            where_conditions.append("contact_phone LIKE ?")
            params.append(f'%{search_phone}%')
        
        where_clause = ' AND '.join(where_conditions) if where_conditions else '1=1'
        
        # 查询所有数据
        cur.execute(f'''
            SELECT supplier_code, supplier_name, contact_person, contact_phone,
                   address, remarks, create_time
            FROM suppliers 
            WHERE {where_clause}
            ORDER BY create_time DESC
        ''', params)
        
        rows = cur.fetchall()
        conn.close()
        
        # 创建Excel
        import pandas as pd
        df = pd.DataFrame(rows, columns=[
            '供应商编号', '供应商名称', '联系人', '联系电话', '地址', '备注', '创建时间'
        ])
        
        # 输出到内存
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='供应商列表')
        
        output.seek(0)
        
        # 生成文件名
        filename = f'供应商列表_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        from flask import send_file
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logger.error(f"导出供应商数据失败: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500

# ========== 销售订单管理 API ==========

@app.route('/sales_orders')
@login_required
def sales_orders_page():
    """销售订单管理页面"""
    return render_template('sales_orders.html')


@app.route('/api/sales_orders', methods=['GET'])
@login_required
def get_sales_orders():
    """获取销售订单列表（分页、搜索）"""
    try:
        page = int(request.args.get('page', 1))
        page_size = int(request.args.get('page_size', 15))
        search_order_code = request.args.get('order_code', '').strip()
        search_customer = request.args.get('customer', '').strip()
        search_status = request.args.get('status', '').strip()
        
        conn = get_db()
        cur = conn.cursor()
        
        # 构建查询条件
        where_conditions = []
        params = []
        
        if search_order_code:
            where_conditions.append("order_code LIKE ?")
            params.append(f'%{search_order_code}%')
        
        if search_customer:
            where_conditions.append("customer_name LIKE ?")
            params.append(f'%{search_customer}%')
        
        if search_status:
            where_conditions.append("status = ?")
            params.append(search_status)
        
        where_clause = ' AND '.join(where_conditions) if where_conditions else '1=1'
        
        # 查询总数
        cur.execute(f"SELECT COUNT(*) FROM sales_orders WHERE {where_clause}", params)
        total = cur.fetchone()[0]
        
        # 查询数据
        offset = (page - 1) * page_size
        cur.execute(f'''
            SELECT id, order_code, customer_id, customer_name, order_date, 
                   delivery_date, total_amount, status, remarks, create_time
            FROM sales_orders 
            WHERE {where_clause}
            ORDER BY create_time DESC
            LIMIT ? OFFSET ?
        ''', params + [page_size, offset])
        
        columns = [desc[0] for desc in cur.description]
        items = [dict(zip(columns, row)) for row in cur.fetchall()]
        
        conn.close()
        
        return jsonify({
        'success': True,
        'data': items,
        'total': total,
        'pagination': {
            'page': page,
            'page_size': page_size,
            'total_pages': (total + page_size - 1) // page_size
        }
    })
        
    except Exception as e:
        logger.error(f"获取销售订单列表失败: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/sales_orders/generate_code', methods=['GET'])
@login_required
def generate_sales_order_code_api():
    """生成新的销售订单编号"""
    try:
        code = generate_sales_order_code()
        return jsonify({'success': True, 'code': code})
    except Exception as e:
        logger.error(f"生成销售订单编号失败: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/sales_orders/<int:order_id>', methods=['GET'])
@login_required
def get_sales_order(order_id):
    """获取单个销售订单详情（含明细）"""
    try:
        conn = get_db()
        cur = conn.cursor()
        
        # 查询订单主表
        cur.execute('''
            SELECT id, order_code, customer_id, customer_name, order_date,
                   delivery_date, total_amount, status, remarks, create_time
            FROM sales_orders WHERE id = ?
        ''', (order_id,))
        
        row = cur.fetchone()
        
        if not row:
            conn.close()
            return jsonify({'success': False, 'message': '订单不存在'}), 404
        
        columns = [desc[0] for desc in cur.description]
        order = dict(zip(columns, row))
        
        # 查询订单明细
        cur.execute('''
            SELECT id, product_name, category, specification, unit,
                   quantity, price, amount, remarks
            FROM sales_order_items
            WHERE order_id = ?
            ORDER BY id
        ''', (order_id,))
        
        item_columns = [desc[0] for desc in cur.description]
        items = [dict(zip(item_columns, row)) for row in cur.fetchall()]
        
        order['items'] = items
        
        conn.close()
        
        return jsonify({'success': True, 'data': order})
        
    except Exception as e:
        logger.error(f"获取销售订单详情失败: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/sales_orders', methods=['POST'])
@login_required
def add_sales_order():
    """添加销售订单"""
    try:
        data = request.get_json()
        
        # 验证必填字段
        if not data.get('customer_id'):
            return jsonify({'success': False, 'message': '请选择客户'}), 400
        
        if not data.get('order_date'):
            return jsonify({'success': False, 'message': '请选择下单日期'}), 400
        
        items = data.get('items', [])
        if not items:
            return jsonify({'success': False, 'message': '请至少添加一条订单明细'}), 400
        
        # 生成订单编号
        if not data.get('order_code'):
            data['order_code'] = generate_sales_order_code()
        
        conn = get_db()
        cur = conn.cursor()
        
        # 获取客户名称
        cur.execute('SELECT customer_name FROM customers WHERE id = ?', (data['customer_id'],))
        customer = cur.fetchone()
        if not customer:
            conn.close()
            return jsonify({'success': False, 'message': '客户不存在'}), 404
        
        customer_name = customer[0]
        
        # 计算订单总额
        total_amount = sum(float(item['amount']) for item in items)
        
        # 插入订单主表
        cur.execute('''
            INSERT INTO sales_orders (
                order_code, customer_id, customer_name, order_date,
                delivery_date, total_amount, status, remarks, create_user
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            data['order_code'],
            data['customer_id'],
            customer_name,
            data['order_date'],
            data.get('delivery_date'),
            total_amount,
            data.get('order_status', '待确认'),
            data.get('remarks'),
            session.get('user', 'system')
        ))
        
        order_id = cur.lastrowid
        
        # 插入订单明细
        for item in items:
            cur.execute('''
                INSERT INTO sales_order_items (
                    order_id, product_name, category, specification,
                    unit, quantity, price, amount, remarks
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                order_id,
                item['product_name'],
                item.get('category'),
                item.get('specification'),
                item.get('unit', '件'),
                item['quantity'],
                item['price'],
                item['amount'],
                item.get('remarks')
            ))
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': '订单创建成功',
            'order_id': order_id,
            'order_code': data['order_code']
        })
        
    except Exception as e:
        logger.error(f"添加销售订单失败: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/sales_orders/<int:order_id>', methods=['PUT'])
@login_required
def update_sales_order(order_id):
    """更新销售订单"""
    try:
        data = request.get_json()
        
        if not data.get('customer_id'):
            return jsonify({'success': False, 'message': '请选择客户'}), 400
        
        if not data.get('order_date'):
            return jsonify({'success': False, 'message': '请选择下单日期'}), 400
        
        items = data.get('items', [])
        if not items:
            return jsonify({'success': False, 'message': '请至少添加一条订单明细'}), 400
        
        conn = get_db()
        cur = conn.cursor()
        
        # 检查订单是否存在
        cur.execute('SELECT id FROM sales_orders WHERE id = ?', (order_id,))
        if not cur.fetchone():
            conn.close()
            return jsonify({'success': False, 'message': '订单不存在'}), 404
        
        # 获取客户名称
        cur.execute('SELECT customer_name FROM customers WHERE id = ?', (data['customer_id'],))
        customer = cur.fetchone()
        if not customer:
            conn.close()
            return jsonify({'success': False, 'message': '客户不存在'}), 404
        
        customer_name = customer[0]
        
        # 计算订单总额
        total_amount = sum(float(item['amount']) for item in items)
        
        # 更新订单主表
        cur.execute('''
            UPDATE sales_orders 
            SET customer_id = ?, customer_name = ?, order_date = ?,
                delivery_date = ?, total_amount = ?, status = ?,
                remarks = ?, update_time = CURRENT_TIMESTAMP
            WHERE id = ?
        ''', (
            data['customer_id'],
            customer_name,
            data['order_date'],
            data.get('delivery_date'),
            total_amount,
            data.get('order_status', '待确认'),
            data.get('remarks'),
            order_id
        ))
        
        # 删除旧明细
        cur.execute('DELETE FROM sales_order_items WHERE order_id = ?', (order_id,))
        
        # 插入新明细
        for item in items:
            cur.execute('''
                INSERT INTO sales_order_items (
                    order_id, product_name, category, specification,
                    unit, quantity, price, amount, remarks
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                order_id,
                item['product_name'],
                item.get('category'),
                item.get('specification'),
                item.get('unit', '件'),
                item['quantity'],
                item['price'],
                item['amount'],
                item.get('remarks')
            ))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': '订单更新成功'})
        
    except Exception as e:
        logger.error(f"更新销售订单失败: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/sales_orders/<int:order_id>', methods=['DELETE'])
@login_required
def delete_sales_order(order_id):
    """删除销售订单"""
    try:
        conn = get_db()
        cur = conn.cursor()
        
        # 检查订单状态
        cur.execute('SELECT status FROM sales_orders WHERE id = ?', (order_id,))
        order = cur.fetchone()
        
        if not order:
            conn.close()
            return jsonify({'success': False, 'message': '订单不存在'}), 404
        
        # 如果订单已发货或已完成，不允许删除
        if order[0] in ['已发货', '已完成']:
            conn.close()
            return jsonify({
                'success': False,
                'message': f'订单状态为"{order[0]}"，不允许删除'
            }), 400
        
        # 删除订单明细（CASCADE会自动删除）
        cur.execute('DELETE FROM sales_order_items WHERE order_id = ?', (order_id,))
        
        # 删除订单
        cur.execute('DELETE FROM sales_orders WHERE id = ?', (order_id,))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': '删除成功'})
        
    except Exception as e:
        logger.error(f"删除销售订单失败: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/sales_orders/export', methods=['GET'])
@login_required
def export_sales_orders():
    """导出销售订单数据到Excel"""
    try:
        import io
        from datetime import datetime
        
        search_order_code = request.args.get('order_code', '').strip()
        search_customer = request.args.get('customer', '').strip()
        search_status = request.args.get('status', '').strip()
        
        conn = get_db()
        cur = conn.cursor()
        
        # 构建查询条件
        where_conditions = []
        params = []
        
        if search_order_code:
            where_conditions.append("o.order_code LIKE ?")
            params.append(f'%{search_order_code}%')
        
        if search_customer:
            where_conditions.append("o.customer_name LIKE ?")
            params.append(f'%{search_customer}%')
        
        if search_status:
            where_conditions.append("o.status = ?")
            params.append(search_status)
        
        where_clause = ' AND '.join(where_conditions) if where_conditions else '1=1'
        
        # 查询订单及明细数据
        cur.execute(f'''
            SELECT 
                o.order_code, o.customer_name, o.order_date, o.delivery_date,
                o.status, i.product_name, i.category, i.specification,
                i.quantity, i.unit, i.price, i.amount, o.remarks
            FROM sales_orders o
            LEFT JOIN sales_order_items i ON o.id = i.order_id
            WHERE {where_clause}
            ORDER BY o.create_time DESC, i.id
        ''', params)
        
        rows = cur.fetchall()
        conn.close()
        
        # 创建Excel
        import pandas as pd
        df = pd.DataFrame(rows, columns=[
            '订单编号', '客户名称', '下单日期', '交货日期', '订单状态',
            '商品名称', '类别', '规格', '数量', '单位', '单价', '金额', '备注'
        ])
        
        # 输出到内存
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='销售订单')
            
            # 获取工作表并设置格式
            worksheet = writer.sheets['销售订单']
            
            # 设置列宽
            worksheet.column_dimensions['A'].width = 15  # 订单编号
            worksheet.column_dimensions['B'].width = 20  # 客户名称
            worksheet.column_dimensions['F'].width = 25  # 商品名称
        
        output.seek(0)
        
        # 生成文件名
        filename = f'销售订单_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        from flask import send_file
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logger.error(f"导出销售订单数据失败: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500
    
    # ========== 采购订单管理 API ==========

def generate_purchase_order_code():
    """生成采购单号"""
    conn = get_db()
    cur = conn.cursor()
    today = datetime.now().strftime('%Y%m%d')
    prefix = f'CG{today}'
    
    cur.execute('''
        SELECT order_code FROM purchase_orders 
        WHERE order_code LIKE ? 
        ORDER BY order_code DESC LIMIT 1
    ''', (f'{prefix}%',))
    
    result = cur.fetchone()
    conn.close()
    
    if result:
        last_num = int(result[0][-4:])
        new_num = last_num + 1
    else:
        new_num = 1
    
    return f'{prefix}{new_num:04d}'


@app.route('/purchase_orders')
@login_required
def purchase_orders_page():
    """采购订单管理页面"""
    return render_template('purchase_orders.html')


@app.route('/api/purchase_orders', methods=['GET'])
@login_required
def get_purchase_orders():
    """获取采购订单列表"""
    try:
        page = int(request.args.get('page', 1))
        page_size = int(request.args.get('page_size', 15))
        search_order_code = request.args.get('order_code', '').strip()
        search_supplier = request.args.get('supplier', '').strip()
        search_status = request.args.get('status', '').strip()
        
        conn = get_db()
        cur = conn.cursor()
        
        where_conditions = []
        params = []
        
        if search_order_code:
            where_conditions.append("order_code LIKE ?")
            params.append(f'%{search_order_code}%')
        
        if search_supplier:
            where_conditions.append("supplier_name LIKE ?")
            params.append(f'%{search_supplier}%')
        
        if search_status:
            where_conditions.append("status = ?")
            params.append(search_status)
        
        where_clause = ' AND '.join(where_conditions) if where_conditions else '1=1'
        
        cur.execute(f"SELECT COUNT(*) FROM purchase_orders WHERE {where_clause}", params)
        total = cur.fetchone()[0]
        
        offset = (page - 1) * page_size
        cur.execute(f'''
            SELECT id, order_code, supplier_id, supplier_name, order_date, 
                   expected_date, total_amount, status, remarks, create_time
            FROM purchase_orders 
            WHERE {where_clause}
            ORDER BY create_time DESC
            LIMIT ? OFFSET ?
        ''', params + [page_size, offset])
        
        columns = [desc[0] for desc in cur.description]
        items = [dict(zip(columns, row)) for row in cur.fetchall()]
        
        conn.close()
        
        return jsonify({
        'success': True,
        'data': items,
        'total': total,
        'pagination': {
            'page': page,
            'page_size': page_size,
            'total_pages': (total + page_size - 1) // page_size
        }
    })
        
    except Exception as e:
        logger.error(f"获取采购订单失败: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/purchase_orders/generate_code', methods=['GET'])
@login_required
def generate_purchase_order_code_api():
    """生成采购单号API"""
    try:
        code = generate_purchase_order_code()
        return jsonify({'success': True, 'code': code})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/purchase_orders/<int:order_id>', methods=['GET'])
@login_required
def get_purchase_order(order_id):
    """获取采购单详情"""
    try:
        conn = get_db()
        cur = conn.cursor()
        
        cur.execute('''
            SELECT id, order_code, supplier_id, supplier_name, order_date,
                   expected_date, total_amount, status, remarks, create_time
            FROM purchase_orders WHERE id = ?
        ''', (order_id,))
        
        row = cur.fetchone()
        if not row:
            conn.close()
            return jsonify({'success': False, 'message': '采购单不存在'}), 404
        
        columns = [desc[0] for desc in cur.description]
        order = dict(zip(columns, row))
        
        cur.execute('''
            SELECT id, product_name, category, specification, unit,
                   quantity, price, amount, remarks
            FROM purchase_order_items
            WHERE order_id = ?
            ORDER BY id
        ''', (order_id,))
        
        item_columns = [desc[0] for desc in cur.description]
        items = [dict(zip(item_columns, row)) for row in cur.fetchall()]
        
        order['items'] = items
        conn.close()
        
        return jsonify({'success': True, 'data': order})
        
    except Exception as e:
        logger.error(f"获取采购单详情失败: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/purchase_orders', methods=['POST'])
@login_required
def add_purchase_order():
    """添加采购订单"""
    try:
        data = request.get_json()
        
        if not data.get('supplier_id'):
            return jsonify({'success': False, 'message': '请选择供应商'}), 400
        
        if not data.get('order_date'):
            return jsonify({'success': False, 'message': '请选择采购日期'}), 400
        
        items = data.get('items', [])
        if not items:
            return jsonify({'success': False, 'message': '请至少添加一条采购明细'}), 400
        
        if not data.get('order_code'):
            data['order_code'] = generate_purchase_order_code()
        
        conn = get_db()
        cur = conn.cursor()
        
        cur.execute('SELECT supplier_name FROM suppliers WHERE id = ?', (data['supplier_id'],))
        supplier = cur.fetchone()
        if not supplier:
            conn.close()
            return jsonify({'success': False, 'message': '供应商不存在'}), 404
        
        supplier_name = supplier[0]
        total_amount = sum(float(item['amount']) for item in items)
        
        cur.execute('''
            INSERT INTO purchase_orders (
                order_code, supplier_id, supplier_name, order_date,
                expected_date, total_amount, status, remarks, create_user
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            data['order_code'],
            data['supplier_id'],
            supplier_name,
            data['order_date'],
            data.get('expected_date'),
            total_amount,
            data.get('order_status', '待确认'),
            data.get('remarks'),
            session.get('user', 'system')
        ))
        
        order_id = cur.lastrowid
        
        for item in items:
            cur.execute('''
                INSERT INTO purchase_order_items (
                    order_id, product_name, category, specification,
                    unit, quantity, price, amount, remarks
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                order_id,
                item['product_name'],
                item.get('category'),
                item.get('specification'),
                item.get('unit', '件'),
                item['quantity'],
                item['price'],
                item['amount'],
                item.get('remarks')
            ))
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': '采购单创建成功',
            'order_id': order_id,
            'order_code': data['order_code']
        })
        
    except Exception as e:
        logger.error(f"添加采购单失败: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/purchase_orders/<int:order_id>', methods=['PUT'])
@login_required
def update_purchase_order(order_id):
    """更新采购订单"""
    try:
        data = request.get_json()
        
        if not data.get('supplier_id'):
            return jsonify({'success': False, 'message': '请选择供应商'}), 400
        
        items = data.get('items', [])
        if not items:
            return jsonify({'success': False, 'message': '请至少添加一条采购明细'}), 400
        
        conn = get_db()
        cur = conn.cursor()
        
        cur.execute('SELECT status FROM purchase_orders WHERE id = ?', (order_id,))
        order = cur.fetchone()
        if not order:
            conn.close()
            return jsonify({'success': False, 'message': '采购单不存在'}), 404
        
        if order[0] in ['已完成', '已取消']:
            conn.close()
            return jsonify({'success': False, 'message': f'订单状态为"{order[0]}"，不允许修改'}), 400
        
        cur.execute('SELECT supplier_name FROM suppliers WHERE id = ?', (data['supplier_id'],))
        supplier = cur.fetchone()
        if not supplier:
            conn.close()
            return jsonify({'success': False, 'message': '供应商不存在'}), 404
        
        supplier_name = supplier[0]
        total_amount = sum(float(item['amount']) for item in items)
        
        cur.execute('''
            UPDATE purchase_orders 
            SET supplier_id = ?, supplier_name = ?, order_date = ?,
                expected_date = ?, total_amount = ?, status = ?,
                remarks = ?, update_time = CURRENT_TIMESTAMP
            WHERE id = ?
        ''', (
            data['supplier_id'],
            supplier_name,
            data['order_date'],
            data.get('expected_date'),
            total_amount,
            data.get('order_status', '待确认'),
            data.get('remarks'),
            order_id
        ))
        
        cur.execute('DELETE FROM purchase_order_items WHERE order_id = ?', (order_id,))
        
        for item in items:
            cur.execute('''
                INSERT INTO purchase_order_items (
                    order_id, product_name, category, specification,
                    unit, quantity, price, amount, remarks
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                order_id,
                item['product_name'],
                item.get('category'),
                item.get('specification'),
                item.get('unit', '件'),
                item['quantity'],
                item['price'],
                item['amount'],
                item.get('remarks')
            ))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': '采购单更新成功'})
        
    except Exception as e:
        logger.error(f"更新采购单失败: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/purchase_orders/<int:order_id>', methods=['DELETE'])
@login_required
def delete_purchase_order(order_id):
    """删除采购订单"""
    try:
        conn = get_db()
        cur = conn.cursor()
        
        cur.execute('SELECT status FROM purchase_orders WHERE id = ?', (order_id,))
        order = cur.fetchone()
        
        if not order:
            conn.close()
            return jsonify({'success': False, 'message': '采购单不存在'}), 404
        
        if order[0] not in ['待确认', '已取消']:
            conn.close()
            return jsonify({'success': False, 'message': f'订单状态为"{order[0]}"，不允许删除'}), 400
        
        cur.execute('DELETE FROM purchase_order_items WHERE order_id = ?', (order_id,))
        cur.execute('DELETE FROM purchase_orders WHERE id = ?', (order_id,))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': '删除成功'})
        
    except Exception as e:
        logger.error(f"删除采购单失败: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/purchase_orders/export', methods=['GET'])
@login_required
def export_purchase_orders():
    """导出采购订单"""
    try:
        import io
        from datetime import datetime
        
        search_order_code = request.args.get('order_code', '').strip()
        search_supplier = request.args.get('supplier', '').strip()
        search_status = request.args.get('status', '').strip()
        
        conn = get_db()
        cur = conn.cursor()
        
        where_conditions = []
        params = []
        
        if search_order_code:
            where_conditions.append("o.order_code LIKE ?")
            params.append(f'%{search_order_code}%')
        
        if search_supplier:
            where_conditions.append("o.supplier_name LIKE ?")
            params.append(f'%{search_supplier}%')
        
        if search_status:
            where_conditions.append("o.status = ?")
            params.append(search_status)
        
        where_clause = ' AND '.join(where_conditions) if where_conditions else '1=1'
        
        cur.execute(f'''
            SELECT 
                o.order_code, o.supplier_name, o.order_date, o.expected_date,
                o.status, i.product_name, i.category, i.specification,
                i.quantity, i.unit, i.price, i.amount, o.remarks
            FROM purchase_orders o
            LEFT JOIN purchase_order_items i ON o.id = i.order_id
            WHERE {where_clause}
            ORDER BY o.create_time DESC, i.id
        ''', params)
        
        rows = cur.fetchall()
        conn.close()
        
        import pandas as pd
        df = pd.DataFrame(rows, columns=[
            '采购单号', '供应商名称', '采购日期', '预计到货', '订单状态',
            '商品名称', '类别', '规格', '数量', '单位', '单价', '金额', '备注'
        ])
        
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='采购订单')
            worksheet = writer.sheets['采购订单']
            worksheet.column_dimensions['A'].width = 18
            worksheet.column_dimensions['B'].width = 20
            worksheet.column_dimensions['F'].width = 25
        
        output.seek(0)
        filename = f'采购订单_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        from flask import send_file
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logger.error(f"导出采购订单失败: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/sales_orders/<int:order_id>/print', methods=['GET'])
@login_required
def print_sales_order(order_id):
    """生成销售订单打印页面"""
    try:
        conn = get_db()
        cur = conn.cursor()
        
        # 查询订单信息
        cur.execute('''
            SELECT o.*, c.contact_person, c.contact_phone, c.address
            FROM sales_orders o
            LEFT JOIN customers c ON o.customer_id = c.id
            WHERE o.id = ?
        ''', (order_id,))
        
        order = dict(cur.fetchone())
        
        # 查询订单明细
        cur.execute('''
            SELECT * FROM sales_order_items WHERE order_id = ? ORDER BY id
        ''', (order_id,))
        
        items = [dict(row) for row in cur.fetchall()]
        
        conn.close()
        
        return render_template('print_sales_order.html', order=order, items=items)
        
    except Exception as e:
        logger.error(f"生成打印页面失败: {str(e)}")
        return f"生成打印页面失败: {str(e)}", 500

@app.route('/api/products', methods=['GET'])
def api_products():
    query = request.args.get('q', '')
    limit = int(request.args.get('limit', 10))
    
    conn = get_db()
    cur = conn.cursor()
    
    if query:
        cur.execute('SELECT id, name FROM products WHERE name LIKE ? ORDER BY name LIMIT ?', 
                    (f'%{query}%', limit))
    else:
        cur.execute('SELECT id, name FROM products ORDER BY name LIMIT ?', (limit,))
        
    products = [{'id': row[0], 'name': row[1]} for row in cur.fetchall()]
    conn.close()
    
    return jsonify(products)

@app.route('/api/product/<int:product_id>', methods=['GET'])
def api_product(product_id):
    conn = get_db()
    cur = conn.cursor()
    
    cur.execute('SELECT * FROM products WHERE id = ?', (product_id,))
    product = cur.fetchone()
    
    if not product:
        return jsonify({'error': '产品不存在'}), 404
    
    result = dict(product)
    conn.close()
    
    return jsonify(result)

@app.route('/api/upload', methods=['POST'])
@login_required
def api_upload():
    if 'file' not in request.files:
        return jsonify({'error': '未选择文件'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': '未选择文件'}), 400
    
    try:
        # 保存临时文件
        filename = secure_filename(file.filename)
        temp_id = str(uuid.uuid4())
        temp_path = os.path.join(UPLOAD_FOLDER, f"{temp_id}_{filename}")
        file.save(temp_path)
        
        # 读取文件内容
        if pd is None:
            return jsonify({'error': '系统缺少pandas库，无法处理Excel/CSV文件'}), 500
            
        if filename.lower().endswith('.csv'):
            df = pd.read_csv(temp_path, encoding='utf-8')
        else:
            df = pd.read_excel(temp_path)
            
        # 准备返回数据
        preview_data = df.head(5).to_dict('records')
        columns = df.columns.tolist()
        
        return jsonify({
            'temp_id': temp_id,
            'filename': filename,
            'columns': columns,
            'preview_data': preview_data
        })
                              
    except Exception as e:
        return jsonify({'error': f'处理文件错误: {str(e)}'}), 500

@app.route('/api/import/upload', methods=['POST'])
@login_required
def api_import_upload():
    try:
        if 'file' not in request.files:
            return jsonify({'error': '未选择文件'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': '未选择文件'}), 400
        
        # 检查pandas
        if pd is None:
            return jsonify({'error': '系统缺少pandas库，无法处理Excel/CSV文件'}), 500
        
        # 保存临时文件
        filename = secure_filename(file.filename)
        temp_id = str(uuid.uuid4())
        temp_path = os.path.join(UPLOAD_FOLDER, f"{temp_id}_{filename}")
        file.save(temp_path)
        
        # 读取文件内容
        if filename.lower().endswith('.csv'):
            df = pd.read_csv(temp_path, encoding='utf-8')
        else:
            df = pd.read_excel(temp_path)
            
        # 准备返回数据
        preview_data = df.head(5).to_dict('records')
        columns = df.columns.tolist()
        
        return jsonify({
            'success': True,
            'temp_id': temp_id,
            'filename': filename,
            'columns': columns,
            'preview_data': preview_data
        })
                              
    except Exception as e:
        logger.exception("API上传错误")
        return jsonify({'error': f'处理文件错误: {str(e)}'}), 500

@app.route('/api/import/task', methods=['POST'])
def api_import_task():
    data = request.json
    if not data or 'temp_id' not in data or 'mapping' not in data:
        return jsonify({'error': '缺少必要参数'}), 400
    
    temp_id = data['temp_id']
    mapping = data['mapping']
    conflict_mode = data.get('conflict_mode', 'skip')
    
    # 查找上传的临时文件
    filepath = _locate_temp_file(temp_id)
    if not filepath:
        return jsonify({'error': '临时文件不存在或已过期'}), 404
    
    # 创建导入任务
    task_id = str(uuid.uuid4())
    filename = os.path.basename(filepath).replace(f"{temp_id}_", "", 1)
    
    conn = get_db()
    cur = conn.cursor()
    now = datetime.now().isoformat()
    
    cur.execute('''
        INSERT INTO import_tasks (id, temp_id, filename, mapping, conflict_mode, 
                                status, created_at, updated_at, total, success, failed)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (task_id, temp_id, filename, json.dumps(mapping), conflict_mode, 
         'pending', now, now, 0, 0, 0))
    conn.commit()
    conn.close()
    
    # 在后台执行导入任务
    _IMPORT_EXECUTOR.submit(_process_import_task, task_id, filepath, mapping, conflict_mode)
    
    return jsonify({
        'success': True,
        'task_id': task_id
    })

@app.route('/api/import/status/<task_id>', methods=['GET'])
def api_import_status(task_id):
    conn = get_db()
    cur = conn.cursor()
    
    cur.execute('SELECT * FROM import_tasks WHERE id = ?', (task_id,))
    task = cur.fetchone()
    
    if not task:
        conn.close()
        return jsonify({'error': '任务不存在'}), 404
    
    # 获取错误信息
    cur.execute('SELECT * FROM import_errors WHERE task_id = ? ORDER BY row_no LIMIT 100', (task_id,))
    errors = [dict(row) for row in cur.fetchall()]
    
    result = dict(task)
    result['errors'] = errors
    
    conn.close()
    return jsonify(result)

# 添加新的API路由用于处理映射
@app.route('/api/import/map', methods=['POST'])
@login_required
def api_import_map():
    data = request.json
    if not data or 'temp_id' not in data or 'mapping' not in data:
        return jsonify({'error': '缺少必要参数'}), 400
    
    temp_id = data['temp_id']
    mapping = data['mapping']
    global_month = data.get('global_month', '')
    conflict_mode = data.get('conflict_mode', 'skip')
    
    # 查找临时文件
    filepath = _locate_temp_file(temp_id)
    if not filepath:
        return jsonify({'error': '临时文件不存在或已过期'}), 404
        
    # 读取文件并进行预处理
    try:
        import pandas as pd
        if filepath.endswith('.csv'):
            df = pd.read_csv(filepath)
        else:
            df = pd.read_excel(filepath)
            
        # 获取必要的映射
        product_col = mapping.get('product')
        if not product_col:
            return jsonify({'error': '必须指定产品列'}), 400
            
        price_cols = mapping.get('price_cols', [])
        if not price_cols:
            return jsonify({'error': '必须至少指定一个价格列'}), 400
            
        # 执行预检
        results = []
        conflicts = []
        
        # 最多检查50行，避免处理时间过长
        for idx, row in df.head(50).iterrows():
            product_name = str(row.get(product_col, ''))
            if not product_name:
                continue
                
            # 归一化产品名并查找
            normalized_name = normalize_product_name(product_name)
            matches = fuzzy_match_product(normalized_name)
            
            row_result = {
                'row': idx + 1,
                'product': product_name,
                'normalized': normalized_name,
                'matches': matches,
                'prices': []
            }
            
            # 检查每个价格列
            for price_info in price_cols:
                price_col = price_info.get('column')
                company = price_info.get('company', '')
                
                if not price_col or not company:
                    continue
                    
                price_val = _parse_number(row.get(price_col))
                
                if price_val is not None:
                    # 检查冲突
                    if matches and conflict_mode != 'overwrite':
                        product_id = matches[0][0]
                        date_val = to_bid_month(row.get(mapping.get('date', '')), global_month)
                        
                        # 查询是否已存在该产品+公司+月份的记录
                        conn = get_db()
                        cur = conn.cursor()
                        cur.execute(
                            "SELECT id, price FROM price_meta WHERE product_id=? AND company=? AND bid_month=?",
                            (product_id, company, date_val)
                        )
                        existing = cur.fetchone()
                        conn.close()
                        
                        if existing:
                            conflicts.append({
                                'row': idx + 1,
                                'product': product_name,
                                'company': company,
                                'existing_price': existing[1],
                                'new_price': price_val
                            })
                    
                    row_result['prices'].append({
                        'company': company,
                        'value': price_val
                    })
            
            results.append(row_result)
        
        return jsonify({
            'success': True,
            'preview': results,
            'conflicts': conflicts,
            'total_rows': len(df)
        })
        
    except Exception as e:
        logger.exception("预检查错误")
        return jsonify({'error': str(e)}), 500

@app.route('/api/import/execute', methods=['POST'])
@login_required
def api_import_execute():
    data = request.json
    if not data or 'temp_id' not in data or 'mapping' not in data:
        return jsonify({'error': '缺少必要参数'}), 400
    
    temp_id = data['temp_id']
    mapping = data['mapping']
    global_month = data.get('global_month', '')
    conflict_mode = data.get('conflict_mode', 'skip')
    
    # 创建导入任务
    task_id = str(uuid.uuid4())
    
    # 记录任务信息
    conn = get_db()
    cur = conn.cursor()
    now = datetime.now().isoformat()
    
    try:
        cur.execute(
            "INSERT INTO import_tasks (id, temp_id, filename, mapping, conflict_mode, status, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            (task_id, temp_id, os.path.basename(_locate_temp_file(temp_id)), json.dumps(mapping), conflict_mode, 'pending', now, now)
        )
        conn.commit()
    except Exception as e:
        conn.close()
        return jsonify({'error': f'创建导入任务失败: {str(e)}'}), 500
    finally:
        conn.close()
    
    # 将任务提交到后台执行
    _IMPORT_EXECUTOR.submit(
        _process_quote_import,
        task_id, 
        temp_id, 
        mapping, 
        global_month, 
        conflict_mode,
        session.get('user_id', 'unknown')
    )
    
    return jsonify({
        'success': True,
        'task_id': task_id
    })

def _process_quote_import(task_id, temp_id, mapping, global_month, conflict_mode, user_id):
    """后台处理导入任务"""
    import pandas as pd
    
    filepath = _locate_temp_file(temp_id)
    if not filepath:
        logger.error(f"找不到临时文件: {temp_id}")
        return
    
    conn = get_db()
    cur = conn.cursor()
    
    try:  # 将 try { 改为 try:
        # 更新任务状态
        now = datetime.now().isoformat()
        cur.execute(
            "UPDATE import_tasks SET status=?, updated_at=? WHERE id=?",
            ('processing', now, task_id)
        )
        conn.commit()
        
        # 读取文件
        if filepath.lower().endswith('.csv'):
            df = pd.read_csv(filepath)
        else:
            df = pd.read_excel(filepath)
        
        # 获取映射配置
        product_col = mapping.get('product')
        if not product_col:
            raise ValueError("必须指定产品列")
        
        price_cols = mapping.get('price_cols', [])
        if not price_cols:
            raise ValueError("必须至少指定一个价格列")
        
        total_rows = len(df)
        success_count = 0
        error_count = 0
        
        # 更新任务总数
        cur.execute(
            "UPDATE import_tasks SET total=? WHERE id=?",
            (total_rows, task_id)
        )
        conn.commit()
        
        # 处理每行数据
        for idx, row in df.iterrows():
            try:
                product_name = str(row.get(product_col, ''))
                if not product_name:
                    continue
                
                # 归一化并查找/创建产品
                normalized_name = normalize_product_name(product_name)
                
                cur.execute(
                    "SELECT id FROM products WHERE normalized_name=?",
                    (normalized_name,)
                )
                product_row = cur.fetchone()
                
                if product_row:
                    product_id = product_row[0]
                else:
                    # 创建新产品
                    cur.execute(
                        "INSERT INTO products (name, normalized_name, created_at) VALUES (?, ?, ?)",
                        (product_name, normalized_name, now)
                    )
                    product_id = cur.lastrowid
                
                # 创建quote记录
                source = f"导入_{temp_id}"
                cur.execute(
                    "INSERT INTO quotes (product_id, source, created_at) VALUES (?, ?, ?)",
                    (product_id, source, now)
                )
                quote_id = cur.lastrowid
                
                # 处理价格列
                date_col = mapping.get('date')
                bid_month = to_bid_month(row.get(date_col) if date_col else None, global_month)
                if not bid_month:
                    bid_month = datetime.now().strftime('%Y-%m')
                
                at_least_one_price = False
                for price_info in price_cols:
                    price_col = price_info.get('column')
                    company = price_info.get('company')
                    price_type = price_info.get('price_type', '中标价(默认)')
                    
                    if not price_col or not company:
                        continue
                    
                    price_val = _parse_number(row.get(price_col))
                    if price_val is None:
                        continue
                    
                    at_least_one_price = True
                    
                    # 检查冲突
                    cur.execute(
                        "SELECT id FROM price_meta WHERE product_id=? AND company=? AND bid_month=?",
                        (product_id, company, bid_month)
                    )
                    existing = cur.fetchone()
                    
                    if existing:
                        if conflict_mode == 'skip':
                            continue
                        elif conflict_mode == 'overwrite':
                            cur.execute(
                                "UPDATE price_meta SET price=?, price_type=?, created_at=? WHERE id=?",
                                (price_val, price_type, now, existing[0])
                            )
                        # 其他模式默认为 'skip'
                    else:
                        # 插入新记录
                        cur.execute(
                            "INSERT INTO price_meta (product_id, bid_month, company, price, price_type, created_at) VALUES (?, ?, ?, ?, ?, ?)",
                            (product_id, bid_month, company, price_val, price_type, now)
                        )
                
                if at_least_one_price:
                    success_count += 1
                
                # 定期更新进度
                if (idx + 1) % 100 == 0 or idx + 1 == total_rows:
                    cur.execute(
                        "UPDATE import_tasks SET success=?, failed=?, updated_at=? WHERE id=?",
                        (success_count, error_count, datetime.now().isoformat(), task_id)
                    )
                    conn.commit()
                
            except Exception as e:
                error_count += 1
                logger.exception(f"处理第{idx+1}行时出错")
                try:
                    cur.execute(
                        "INSERT INTO import_errors (task_id, row_no, raw, error_msg) VALUES (?, ?, ?, ?)",
                        (task_id, idx + 1, json.dumps(dict(row)), str(e))
                    )
                except:
                    pass
            
        # 完成导入
        cur.execute(
            "UPDATE import_tasks SET status=?, success=?, failed=?, updated_at=? WHERE id=?",
            ('completed', success_count, error_count, datetime.now().isoformat(), task_id)
        )
        conn.commit()
        
    except Exception as e:
        logger.exception(f"导入任务出错: {str(e)}")
        try:
            cur.execute(
                "UPDATE import_tasks SET status=?, error_msg=?, updated_at=? WHERE id=?",
                ('failed', str(e), datetime.now().isoformat(), task_id)
            )
            conn.commit()
        except:
            pass
    
    finally:
        conn.close()

# 在适当的位置添加此函数，建议放在其他导入任务相关函数附近

def _process_import_task(task_id, filepath, mapping, conflict_mode):
    """处理导入任务的后台函数"""
    import pandas as pd
    
    conn = get_db()
    cur = conn.cursor()
    
    try:
        # 更新任务状态为处理中
        now = datetime.now().isoformat()
        cur.execute(
            "UPDATE import_tasks SET status=?, updated_at=? WHERE id=?",
            ('processing', now, task_id)
        )
        conn.commit()
        
        # 读取文件
        if filepath.lower().endswith('.csv'):
            df = pd.read_csv(filepath)
        else:
            df = pd.read_excel(filepath)
        
        total_rows = len(df)
        success_count = 0
        error_count = 0
        
        # 更新任务总记录数
        cur.execute(
            "UPDATE import_tasks SET total=? WHERE id=?",
            (total_rows, task_id)
        )
        conn.commit()
        
        # 执行导入逻辑（根据mapping参数来处理）
        # 这里实现基本的导入流程，根据您的具体需求可能需要调整
        for idx, row in df.iterrows():
            try:
                # 这里需要根据mapping参数处理每行数据
                # 例如: 根据mapping获取列名，从行中提取数据，插入到相应表中
                
                # 导入成功计数
                success_count += 1
                
                # 定期更新进度
                if (idx + 1) % 100 == 0 or idx + 1 == total_rows:
                    cur.execute(
                        "UPDATE import_tasks SET success=?, failed=?, updated_at=? WHERE id=?",
                        (success_count, error_count, datetime.now().isoformat(), task_id)
                    )
                    conn.commit()
                    
            except Exception as e:
                error_count += 1
                logger.exception(f"处理第{idx+1}行时出错: {str(e)}")
                try:
                    cur.execute(
                        "INSERT INTO import_errors (task_id, row_no, raw, error_msg) VALUES (?, ?, ?, ?)",
                        (task_id, idx + 1, json.dumps(dict(row)), str(e))
                    )
                    conn.commit()
                except:
                    pass
        
        # 完成导入
        cur.execute(
            "UPDATE import_tasks SET status=?, updated_at=? WHERE id=?",
            ('completed', datetime.now().isoformat(), task_id)
        )
        conn.commit()
        
    except Exception as e:
        logger.exception(f"导入任务出错: {str(e)}")
        try:
            cur.execute(
                "UPDATE import_tasks SET status=?, error_msg=?, updated_at=? WHERE id=?",
                ('failed', str(e), datetime.now().isoformat(), task_id)
            )
            conn.commit()
        except:
            pass
    finally:
        conn.close()



@app.route('/api/fields/add', methods=['POST'])
@login_required
def add_field():
    try:
        data = request.get_json()
        table = data.get('table')
        name = data.get('name')
        field_type = data.get('type')
        
        if not all([table, name, field_type]):
            return jsonify({'error': '参数不完整'}), 400
            
        # 验证表名安全性
        if table not in ['quotes', 'orders']:
            return jsonify({'error': '不支持的表名'}), 400
            
        # 放宽字段名验证，支持中文
        if not name.strip() or len(name.strip()) == 0:
            return jsonify({'error': '字段名不能为空'}), 400
            
        # 简单过滤危险字符
        dangerous_chars = [';', '--', '/*', '*/', 'DROP', 'DELETE', 'UPDATE']
        name_upper = name.upper()
        for dangerous in dangerous_chars:
            if dangerous in name_upper:
                return jsonify({'error': f'字段名不能包含危险字符: {dangerous}'}), 400
        
        # 转换字段类型
        sql_type_map = {
            'text': 'TEXT',
            'int': 'INTEGER', 
            'float': 'REAL'
        }
        sql_type = sql_type_map.get(field_type, 'TEXT')
        
        # 执行 ALTER TABLE 添加字段（使用反引号包围字段名）
        conn = get_db()
        cur = conn.cursor()
        cur.execute(f'ALTER TABLE {table} ADD COLUMN `{name}` {sql_type}')
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': f'字段 {name} 添加成功'})
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/calculation/get_prices', methods=['POST'])
@login_required
def api_get_prices():
    """获取产品价格数据用于计算表格"""
    try:
        data = request.get_json()
        products = data.get('products', [])
        company_columns = data.get('company_columns', [])
        
        if not products or not company_columns:
            return jsonify({'error': '缺少必要参数'}), 400
        
        conn = get_db()
        cur = conn.cursor()
        
        # 构建结果数据
        result_data = {}
        
        for product in products:
            result_data[product] = {}
            
            for col_info in company_columns:
                company = col_info.get('company', '')
                year = col_info.get('year', '')
                month = col_info.get('month', '')
                col_name = col_info.get('name', '')
                
                if not all([company, year, month]):
                    result_data[product][col_name] = '参数错误'
                    continue
                
                # 构建日期模式 YYYY-MM%
                date_pattern = f"{year}-{int(month):02d}%"
                
                # 查询数据库
                cur.execute('''
                    SELECT price FROM quotes 
                    WHERE product = ? AND company = ? AND bid_date LIKE ?
                    ORDER BY bid_date DESC LIMIT 1
                ''', (product, company, date_pattern))
                
                row = cur.fetchone()
                
                if row:
                    result_data[product][col_name] = float(row[0])
                else:
                    result_data[product][col_name] = '无数据'
        
        conn.close()
        
        return jsonify({
            'success': True,
            'data': result_data
        })
        
    except Exception as e:
        logger.exception("获取价格数据错误")
        return jsonify({'error': str(e)}), 500

@app.route('/api/calculation/save_result', methods=['POST'])
@login_required
def api_save_calculation_result():
    """保存计算结果"""
    try:
        data = request.get_json()
        table_data = data.get('table_data', [])
        formula_info = data.get('formula_info', {})
        
        if not table_data:
            return jsonify({'error': '没有数据需要保存'}), 400
        
        conn = get_db()
        cur = conn.cursor()
        
        # 保存到一个新的计算结果表（如果需要的话）
        # 这里先简单记录到 settings 表中
        now = datetime.now().isoformat()
        result_data = {
            'table_data': table_data,
            'formula_info': formula_info,
            'created_at': now
        }
        
        cur.execute(
            'INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)',
            (f'calculation_result_{now}', json.dumps(result_data))
        )
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': '计算结果保存成功'
        })
        
    except Exception as e:
        logger.exception("保存计算结果错误")
        return jsonify({'error': str(e)}), 500
    
@app.route('/api/calculation/evaluate_formula', methods=['POST'])
@login_required
def api_evaluate_formula():
    """计算中文公式"""
    try:
        data = request.get_json()
        formula = data.get('formula', '').strip()
        table_data = data.get('table_data', [])
        row_index = data.get('row_index', 0)
        
        if not formula:
            return jsonify({'success': True, 'result': ''})
        
        if not formula.startswith('='):
            return jsonify({'success': True, 'result': formula})
        
        # 移除开头的等号
        expression = formula[1:]
        
        # 中文函数替换为英文
        chinese_functions = {
            '求和': 'SUM',
            '平均值': 'AVERAGE',
            '最大值': 'MAX', 
            '最小值': 'MIN',
            '如果': 'IF',
            '计数': 'COUNT'
        }
        
        for chinese, english in chinese_functions.items():
            expression = expression.replace(chinese, english)
        
        # 解析单元格引用和列名引用
        result = parse_and_calculate_formula(expression, table_data, row_index)
        
        return jsonify({
            'success': True,
            'result': result,
            'original_formula': formula
        })
        
    except Exception as e:
        logger.exception("公式计算错误")
        return jsonify({
            'success': False,
            'error': str(e),
            'result': '#错误#'
        })

def parse_and_calculate_formula(expression, table_data, row_index):
    """增强的公式解析和计算"""
    try:
        # 处理中文函数转换
        expression = convert_chinese_functions(expression)
        
        # 处理条件函数
        expression = process_conditional_functions(expression)
        
        # 处理Excel式引用（如B1, C1等）
        expression = process_excel_references(expression, table_data, row_index)
        
        # 处理数学函数
        expression = process_math_functions(expression, table_data, row_index)
        
        # 最终计算
        result = safe_eval(expression)
        
        return result
        
    except Exception as e:
        logger.exception(f"公式计算错误: {expression}")
        return f'#错误#{str(e)}'

def convert_chinese_functions(expression):
    """转换中文函数为英文函数"""
    # 中文函数映射
    chinese_functions = {
        '如果': 'IF',
        '求和': 'SUM',
        '平均': 'AVERAGE',
        '最大': 'MAX',
        '最小': 'MIN',
        '绝对值': 'ABS',
        '四舍五入': 'ROUND'
    }
    
    for chinese, english in chinese_functions.items():
        expression = expression.replace(chinese, english)
    
    return expression

def process_conditional_functions(expression):
    """处理条件函数 IF(condition, true_value, false_value)"""
    import re
    
    # 查找所有IF函数
    def replace_if_function(match):
        full_match = match.group(0)
        content = match.group(1)
        
        try:
            # 简单解析参数（需要考虑嵌套括号）
            params = parse_function_parameters(content)
            
            if len(params) != 3:
                return f'#错误#IF函数需要3个参数'
            
            condition, true_val, false_val = params
            
            # 构建Python条件表达式
            return f'({true_val} if ({condition}) else {false_val})'
            
        except Exception as e:
            return f'#错误#{str(e)}'
    
    # 替换IF函数
    pattern = r'IF\s*\(\s*([^)]+(?:\([^)]*\)[^)]*)*)\s*\)'
    while re.search(pattern, expression):
        expression = re.sub(pattern, replace_if_function, expression)
    
    return expression

def parse_function_parameters(param_string):
    """解析函数参数，处理嵌套括号和逗号"""
    params = []
    current_param = ""
    paren_level = 0
    
    for char in param_string:
        if char == '(':
            paren_level += 1
        elif char == ')':
            paren_level -= 1
        elif char == ',' and paren_level == 0:
            params.append(current_param.strip())
            current_param = ""
            continue
        
        current_param += char
    
    if current_param.strip():
        params.append(current_param.strip())
    
    return params

def process_excel_references(expression, table_data, row_index):
    """处理Excel式引用 (B1, C1, etc.)"""
    import re
    
    def replace_excel_ref(match):
        col_letter = match.group(1)
        row_num = int(match.group(2))
        
        # 转换列字母为数字 (A=0, B=1, C=2, ...)
        col_index = ord(col_letter.upper()) - ord('A')
        row_idx = row_num - 1  # Excel行从1开始，数组从0开始
        
        try:
            if row_idx < len(table_data) and col_index < len(table_data[row_idx]):
                value = table_data[row_idx][col_index]
                if isinstance(value, (int, float)):
                    return str(value)
                elif value == '无数据':
                    return '0'
                else:
                    try:
                        return str(float(value))
                    except:
                        return '0'
            else:
                return '0'
        except:
            return '0'
    
    # 匹配Excel式引用 (如 B1, C2, AA10等)
    pattern = r'([A-Z]+)(\d+)'
    expression = re.sub(pattern, replace_excel_ref, expression)
    
    return expression

def process_math_functions(expression, table_data, row_index):
    """处理数学函数"""
    import re
    
    # 处理SUM函数
    def replace_sum_function(match):
        content = match.group(1)
        try:
            # 解析范围 (如 B1:D1)
            if ':' in content:
                start_ref, end_ref = content.split(':')
                # 简单实现：假设是同一行的范围
                start_col = ord(start_ref[0]) - ord('A')
                end_col = ord(end_ref[0]) - ord('A')
                
                total = 0
                for col in range(start_col, end_col + 1):
                    if col < len(table_data[row_index]):
                        value = table_data[row_index][col]
                        if isinstance(value, (int, float)):
                            total += value
                        elif value != '无数据':
                            try:
                                total += float(value)
                            except:
                                pass
                
                return str(total)
            else:
                return '0'
        except:
            return '0'
    
    # 替换SUM函数
    expression = re.sub(r'SUM\s*\(\s*([^)]+)\s*\)', replace_sum_function, expression)
    
    # 处理其他数学函数...
    # ABS, ROUND, MAX, MIN等可以类似实现
    
    return expression

def safe_eval(expression):
    """安全的表达式计算"""
    import re
    import math
    
    # 允许的函数和常量
    allowed_names = {
        '__builtins__': {},
        'abs': abs,
        'round': round,
        'max': max,
        'min': min,
        'sum': sum,
        'pow': pow,
        'sqrt': math.sqrt,
        'sin': math.sin,
        'cos': math.cos,
        'tan': math.tan,
        'pi': math.pi,
        'e': math.e
    }
    
    # 检查表达式是否安全
    if re.search(r'[^0-9+\-*/().< >= !=and or not\s]', expression.replace('if', '').replace('else', '')):
        # 如果包含不安全字符，进行更严格的检查
        forbidden_patterns = [
            r'import\s+', r'exec\s*\(', r'eval\s*\(', r'open\s*\(',
            r'file\s*\(', r'input\s*\(', r'raw_input\s*\(',
            r'__.*__', r'\..*\('
        ]
        
        for pattern in forbidden_patterns:
            if re.search(pattern, expression, re.IGNORECASE):
                raise ValueError("不安全的表达式")
    
    try:
        result = eval(expression, allowed_names, {})
        return float(result) if isinstance(result, (int, float)) else result
    except Exception as e:
        raise ValueError(f"计算错误: {str(e)}")
@app.route('/api/calculation/export_excel', methods=['POST'])
@login_required
def api_export_excel():
    """导出计算表格为Excel"""
    try:
        data = request.get_json()
        table_data = data.get('table_data', [])
        headers = data.get('headers', [])
        
        if not table_data or not headers:
            return jsonify({'error': '没有数据需要导出'}), 400
        
        # 创建Excel工作簿
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "计算分析结果"
        
        # 写入表头
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # 写入数据
        for row_idx, row_data in enumerate(table_data, 2):
            for col_idx, cell_value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                
                # 处理不同类型的数据
                if isinstance(cell_value, (int, float)):
                    cell.value = cell_value
                    cell.number_format = '#,##0.00'
                elif cell_value == '无数据':
                    cell.value = '无数据'
                    cell.font = Font(color="999999")
                else:
                    cell.value = str(cell_value)
        
        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 保存到临时文件
        import tempfile
        import os
        from datetime import datetime
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"计算分析结果_{timestamp}.xlsx"
        
        temp_dir = tempfile.gettempdir()
        filepath = os.path.join(temp_dir, filename)
        wb.save(filepath)
        
        # 返回文件下载链接
        return jsonify({
            'success': True,
            'download_url': f'/download_temp/{filename}',
            'filename': filename
        })
        
    except ImportError:
        return jsonify({'error': '请安装 openpyxl 库：pip install openpyxl'}), 500
    except Exception as e:
        logger.exception("导出Excel失败")
        return jsonify({'error': str(e)}), 500

@app.route('/download_temp/<filename>')
@login_required
def download_temp_file(filename):
    """下载临时文件"""
    import tempfile
    import os
    from flask import send_file
    
    temp_dir = tempfile.gettempdir()
    filepath = os.path.join(temp_dir, filename)
    
    if os.path.exists(filepath):
        return send_file(filepath, as_attachment=True, download_name=filename)
    else:
        return "文件不存在", 404
    
    # 在现有的 API 路由部分添加这个新的 API

@app.route('/api/calculation/get_auto_products', methods=['POST'])
@login_required
def api_get_auto_products():
    """根据第一个价格列条件自动获取产品列表"""
    try:
        data = request.get_json()
        first_column = data.get('first_column', {})
        
        company = first_column.get('company', '')
        year = first_column.get('year', '')
        month = first_column.get('month', '')
        
        if not all([company, year, month]):
            return jsonify({'error': '缺少必要参数'}), 400
        
        conn = get_db()
        cur = conn.cursor()
        
        # 构建日期模式 YYYY-MM%
        date_pattern = f"{year}-{int(month):02d}%"
        
        # 查询该公司+年月条件下有价格数据的所有产品
        cur.execute('''
            SELECT DISTINCT product FROM quotes 
            WHERE company = ? AND bid_date LIKE ?
            ORDER BY id
        ''', (company, date_pattern))
        
        rows = cur.fetchall()
        products = [row[0] for row in rows]
        conn.close()
        
        if not products:
            return jsonify({'error': '无数据，请检查第一个价格列的设置'}), 404
        
        # 检查是否有重复产品（理论上不应该，但按需求检查）
        if len(products) != len(set(products)):
            return jsonify({'error': '数据中存在重复产品，请检查数据完整性'}), 400
        
        return jsonify({
            'success': True,
            'products': products
        })
        
    except Exception as e:
        logger.exception("自动获取产品列表错误")
        return jsonify({'error': str(e)}), 500
    
@app.route('/api/calculation/evaluate_column_formula', methods=['POST'])
@login_required
def api_evaluate_column_formula():
    """计算整列公式"""
    try:
        data = request.get_json()
        formula = data.get('formula', '').strip()
        table_data = data.get('table_data', [])
        column_index = data.get('column_index', -1)
        column_headers = data.get('column_headers', [])
        
        if not formula or column_index < 0:
            return jsonify({'error': '缺少必要参数'}), 400
        
        if not formula.startswith('='):
            formula = '=' + formula
        
        # 移除开头的等号
        expression = formula[1:]
        
        # 处理列名引用（优先级最高）
        expression_with_column_refs = replace_column_references(expression, column_headers)
        
        results = []
        
        # 为每一行计算公式
        for row_index in range(len(table_data)):
            try:
                # 使用增强的列名替换
                row_expression = enhance_column_reference_replacement(expression_with_column_refs, table_data, row_index, column_headers)
                
                # 计算公式
                result = parse_and_calculate_formula(row_expression, table_data, row_index)
                results.append(result)
                
            except Exception as e:
                logger.exception(f"计算第{row_index+1}行公式时出错")
                results.append(f'#错误#{str(e)}')
        
        return jsonify({
            'success': True,
            'results': results,
            'formula': formula
        })
        
    except Exception as e:
        logger.exception("列公式计算错误")
        return jsonify({'error': str(e)}), 500

def replace_column_references(expression, column_headers):
    """替换列名引用为列索引标记"""
    import re
    
    # 按长度倒序排列列名，避免短名称被长名称包含时的替换问题
    sorted_headers = sorted(column_headers, key=len, reverse=True)
    
    for i, header in enumerate(sorted_headers):
        if header in expression:
            # 找到该列名在原始列表中的位置
            actual_index = column_headers.index(header)
            # 替换为特殊标记，稍后替换为实际值
            expression = expression.replace(header, f'__COL_{actual_index}__')
    
    return expression

def substitute_column_values(expression, table_data, row_index, column_headers):
    """将列索引标记替换为具体数值"""
    import re
    
    def replace_col_marker(match):
        col_index = int(match.group(1))
        
        if 0 <= row_index < len(table_data) and 0 <= col_index < len(table_data[row_index]):
            value = table_data[row_index][col_index]
            
            # 处理不同类型的值
            if value == '无数据' or value == '获取中...' or value == '获取失败':
                return '0'
            elif isinstance(value, (int, float)):
                return str(value)
            else:
                try:
                    # 尝试转换为数字
                    float_val = float(str(value).replace(',', ''))
                    return str(float_val)
                except:
                    return '0'
        else:
            return '0'
    
    # 替换所有列标记
    result = re.sub(r'__COL_(\d+)__', replace_col_marker, expression)
    
    return result

def enhance_column_reference_replacement(expression, table_data, row_index, column_headers):
    """增强的列引用替换，支持计算列引用"""
    import re
    
    # 先处理列名引用
    expression = substitute_column_values(expression, table_data, row_index, column_headers)
    
    # 处理可能遗留的列名（如果有计算列相互引用）
    for i, header in enumerate(column_headers):
        if header in expression:
            if 0 <= row_index < len(table_data) and i < len(table_data[row_index]):
                value = table_data[row_index][i]
                
                if value == '无数据' or value == '获取中...' or value == '获取失败':
                    value = 0
                elif isinstance(value, str):
                    try:
                        value = float(value.replace(',', ''))
                    except:
                        value = 0
                
                # 使用更精确的替换，避免部分匹配
                expression = re.sub(r'\b' + re.escape(header) + r'\b', str(value), expression)
    
    return expression

# 批量删除
@app.route('/api/smart_quotes/batch_delete', methods=['POST'])
@login_required
def api_batch_delete():
    """批量删除智能报价记录"""
    try:
        data = request.get_json()
        ids = data.get('ids', [])
        
        if not ids:
            return jsonify({'success': False, 'error': '未选择要删除的记录'})
        
        conn = get_db()
        cur = conn.cursor()
        
        # 删除记录
        placeholders = ','.join(['?' for _ in ids])
        cur.execute(f'DELETE FROM quotes WHERE id IN ({placeholders})', ids)
        
        deleted_count = cur.rowcount
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'deleted_count': deleted_count
        })
        
    except Exception as e:
        logger.exception("批量删除错误")
        return jsonify({'success': False, 'error': str(e)}), 500

# 批量修改
@app.route('/api/smart_quotes/batch_update', methods=['POST'])
@login_required
def api_batch_update():
    """批量修改智能报价记录"""
    try:
        data = request.get_json()
        ids = data.get('ids', [])
        fields = data.get('fields', {})
        
        if not ids:
            return jsonify({'success': False, 'error': '未选择要修改的记录'})
        
        if not fields:
            return jsonify({'success': False, 'error': '未指定修改内容'})
        
        conn = get_db()
        cur = conn.cursor()
        
        updated_count = 0
        
        for quote_id in ids:
            updates = []
            params = []
            
            # 处理公司修改
            if 'company' in fields and fields['company']:
                updates.append('company = ?')
                params.append(fields['company'])
            
            # 处理日期修改
            if 'bid_date' in fields and fields['bid_date']:
                updates.append('bid_date = ?')
                params.append(fields['bid_date'])
            
            # 处理价格调整
            if 'price_adjustment' in fields and fields['price_adjustment']:
                adjustment = fields['price_adjustment']
                action = adjustment.get('action')
                value = adjustment.get('value')
                
                if action and value is not None:
                    # 获取当前价格
                    cur.execute('SELECT price FROM quotes WHERE id = ?', (quote_id,))
                    row = cur.fetchone()
                    if row and row[0]:
                        current_price = float(row[0])
                        new_price = current_price
                        
                        if action == 'multiply':
                            new_price = current_price * value
                        elif action == 'add':
                            new_price = current_price + value
                        elif action == 'subtract':
                            new_price = current_price - value
                        elif action == 'set':
                            new_price = value
                        
                        if new_price > 0:  # 确保价格为正
                            updates.append('price = ?')
                            params.append(new_price)
            
            # 执行更新
            if updates:
                params.append(quote_id)
                sql = f'UPDATE quotes SET {", ".join(updates)} WHERE id = ?'
                cur.execute(sql, params)
                updated_count += cur.rowcount
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'updated_count': updated_count
        })
        
    except Exception as e:
        logger.exception("批量修改错误")
        return jsonify({'success': False, 'error': str(e)}), 500

# 批量复制
@app.route('/api/smart_quotes/batch_copy', methods=['POST'])
@login_required
def api_batch_copy():
    """批量复制智能报价记录"""
    try:
        data = request.get_json()
        ids = data.get('ids', [])
        target_company = data.get('target_company')
        target_date = data.get('target_date')
        price_adjustment = data.get('price_adjustment')
        
        if not ids:
            return jsonify({'success': False, 'error': '未选择要复制的记录'})
        
        if not target_company or not target_date:
            return jsonify({'success': False, 'error': '目标公司和日期不能为空'})
        
        conn = get_db()
        cur = conn.cursor()
        
        copied_count = 0
        
        # 获取要复制的记录
        placeholders = ','.join(['?' for _ in ids])
        cur.execute(f'SELECT product, price, qty, remarks FROM quotes WHERE id IN ({placeholders})', ids)
        records = cur.fetchall()
        
        for record in records:
            product, price, qty, remarks = record
            
            # 处理价格调整
            new_price = float(price) if price else 0
            if price_adjustment:
                action = price_adjustment.get('action')
                value = price_adjustment.get('value')
                
                if action and value is not None:
                    if action == 'multiply':
                        new_price = new_price * value
                    elif action == 'add':
                        new_price = new_price + value
                    elif action == 'subtract':
                        new_price = new_price - value
            
            # 检查是否已存在相同记录
            cur.execute(
                'SELECT id FROM quotes WHERE product=? AND company=? AND bid_date=?',
                (product, target_company, target_date)
            )
            existing = cur.fetchone()
            
            if existing:
                # 更新现有记录
                cur.execute(
                    'UPDATE quotes SET price=?, qty=?, remarks=? WHERE id=?',
                    (new_price, qty, f'批量复制_{datetime.now().strftime("%Y%m%d")}', existing[0])
                )
            else:
                # 插入新记录
                cur.execute('''
                    INSERT INTO quotes (product, company, price, qty, bid_date, remarks)
                    VALUES (?, ?, ?, ?, ?, ?)
                ''', (product, target_company, new_price, qty, target_date, 
                     f'批量复制_{datetime.now().strftime("%Y%m%d")}'))
            
            copied_count += 1
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'copied_count': copied_count
        })
        
    except Exception as e:
        logger.exception("批量复制错误")
        return jsonify({'success': False, 'error': str(e)}), 500

# 批量导出
@app.route('/api/smart_quotes/batch_export', methods=['POST'])
@login_required
def api_batch_export():
    """批量导出智能报价记录到Excel"""
    try:
        ids_json = request.form.get('ids')
        if not ids_json:
            return jsonify({'success': False, 'error': '未选择要导出的记录'})
        
        ids = json.loads(ids_json)
        
        conn = get_db()
        cur = conn.cursor()
        
        # 获取选中的记录
        placeholders = ','.join(['?' for _ in ids])
        cur.execute(f'''
            SELECT product, company, price, qty, bid_date, remarks
            FROM quotes 
            WHERE id IN ({placeholders})
            ORDER BY bid_date DESC, product
        ''', ids)
        
        records = cur.fetchall()
        conn.close()
        
        if not records:
            return jsonify({'success': False, 'error': '没有找到要导出的记录'})
        
        # 创建Excel文件
        import io
        output = io.BytesIO()
        
        if pd is not None:
            # 使用pandas创建Excel
            df = pd.DataFrame(records, columns=['产品名称', '中标公司', '中标价格', '预计用量', '中标年月', '备注'])
            df.to_excel(output, index=False, engine='openpyxl')
        else:
            # 简单的CSV格式
            import csv
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(['产品名称', '中标公司', '中标价格', '预计用量', '中标年月', '备注'])
            writer.writerows(records)
        
        output.seek(0)
        
        # 生成文件名
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'智能报价导出_{timestamp}.xlsx' if pd else f'智能报价导出_{timestamp}.csv'
        
        # 返回文件
        from flask import send_file
        return send_file(
            output if pd else io.BytesIO(output.getvalue().encode('utf-8-sig')),
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' if pd else 'text/csv'
        )
        
    except Exception as e:
        logger.exception("批量导出错误")
        return jsonify({'success': False, 'error': str(e)}), 500

# ========== 客户管理 API ==========

@app.route('/customers')
@login_required
def customers_page():
    """客户管理页面"""
    return render_template('customers.html')


@app.route('/api/customers', methods=['GET'])
@login_required
def get_customers():
    """获取客户列表（分页、搜索）"""
    try:
        page = int(request.args.get('page', 1))
        page_size = int(request.args.get('page_size', 15))
        search_name = request.args.get('name', '').strip()
        search_phone = request.args.get('phone', '').strip()
        
        conn = get_db()
        cur = conn.cursor()
        
        # 构建查询条件
        where_conditions = []
        params = []
        
        if search_name:
            where_conditions.append("customer_name LIKE ?")
            params.append(f'%{search_name}%')
        
        if search_phone:
            where_conditions.append("contact_phone LIKE ?")
            params.append(f'%{search_phone}%')
        
        where_clause = ' AND '.join(where_conditions) if where_conditions else '1=1'
        
        # 查询总数
        cur.execute(f"SELECT COUNT(*) FROM customers WHERE {where_clause}", params)
        total = cur.fetchone()[0]
        
        # 查询数据
        offset = (page - 1) * page_size
        cur.execute(f'''
            SELECT id, customer_code, customer_name, contact_person, 
                   contact_phone, address, remarks, create_time, update_time
            FROM customers 
            WHERE {where_clause}
            ORDER BY create_time DESC
            LIMIT ? OFFSET ?
        ''', params + [page_size, offset])
        
        columns = [desc[0] for desc in cur.description]
        items = [dict(zip(columns, row)) for row in cur.fetchall()]
        
        conn.close()
        
           
        return jsonify({
        'success': True,
        'data': items,              # 直接返回数组
        'total': total,             # 移到外层
        'pagination': {             # 可选：分页信息
            'page': page,
            'page_size': page_size,
            'total_pages': (total + page_size - 1) // page_size
        }
    })
        
    except Exception as e:
        logger.error(f"获取客户列表失败: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/customers/generate_code', methods=['GET'])
@login_required
def generate_customer_code_api():
    """生成新的客户编号"""
    try:
        code = generate_customer_code()
        return jsonify({'success': True, 'code': code})
    except Exception as e:
        logger.error(f"生成客户编号失败: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/customers/<int:customer_id>', methods=['GET'])
@login_required
def get_customer(customer_id):
    """获取单个客户详情"""
    try:
        conn = get_db()
        cur = conn.cursor()
        
        cur.execute('''
            SELECT id, customer_code, customer_name, contact_person,
                   contact_phone, address, remarks, create_time, update_time
            FROM customers WHERE id = ?
        ''', (customer_id,))
        
        row = cur.fetchone()
        conn.close()
        
        if not row:
            return jsonify({'success': False, 'message': '客户不存在'}), 404
        
        columns = [desc[0] for desc in cur.description]
        customer = dict(zip(columns, row))
        
        return jsonify({'success': True, 'data': customer})
        
    except Exception as e:
        logger.error(f"获取客户详情失败: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/customers', methods=['POST'])
@login_required
def add_customer():
    """添加客户"""
    try:
        data = request.get_json()
        
        # 验证必填字段
        if not data.get('customer_name'):
            return jsonify({'success': False, 'message': '客户名称不能为空'}), 400
        
        if not data.get('customer_code'):
            data['customer_code'] = generate_customer_code()
        
        conn = get_db()
        cur = conn.cursor()
        
        # 检查编号是否重复
        cur.execute('SELECT id FROM customers WHERE customer_code = ?', (data['customer_code'],))
        if cur.fetchone():
            conn.close()
            return jsonify({'success': False, 'message': '客户编号已存在'}), 400
        
        # 插入数据
        cur.execute('''
            INSERT INTO customers (customer_code, customer_name, contact_person,
                                 contact_phone, address, remarks)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (
            data['customer_code'],
            data['customer_name'],
            data.get('contact_person'),
            data.get('contact_phone'),
            data.get('address'),
            data.get('remarks')
        ))
        
        conn.commit()
        customer_id = cur.lastrowid
        conn.close()
        
        return jsonify({'success': True, 'message': '添加成功', 'id': customer_id})
        
    except Exception as e:
        logger.error(f"添加客户失败: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/customers/<int:customer_id>', methods=['PUT'])
@login_required
def update_customer(customer_id):
    """更新客户"""
    try:
        data = request.get_json()
        
        if not data.get('customer_name'):
            return jsonify({'success': False, 'message': '客户名称不能为空'}), 400
        
        conn = get_db()
        cur = conn.cursor()
        
        # 检查客户是否存在
        cur.execute('SELECT id FROM customers WHERE id = ?', (customer_id,))
        if not cur.fetchone():
            conn.close()
            return jsonify({'success': False, 'message': '客户不存在'}), 404
        
        # 更新数据
        cur.execute('''
            UPDATE customers 
            SET customer_name = ?, contact_person = ?, contact_phone = ?,
                address = ?, remarks = ?, update_time = CURRENT_TIMESTAMP
            WHERE id = ?
        ''', (
            data['customer_name'],
            data.get('contact_person'),
            data.get('contact_phone'),
            data.get('address'),
            data.get('remarks'),
            customer_id
        ))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': '更新成功'})
        
    except Exception as e:
        logger.error(f"更新客户失败: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/customers/<int:customer_id>', methods=['DELETE'])
@login_required
def delete_customer(customer_id):
    """删除客户"""
    try:
        conn = get_db()
        cur = conn.cursor()
        
        # 检查是否有关联订单
        cur.execute('SELECT COUNT(*) FROM sales_orders WHERE customer_id = ?', (customer_id,))
        order_count = cur.fetchone()[0]
        
        if order_count > 0:
            conn.close()
            return jsonify({
                'success': False, 
                'message': f'该客户有 {order_count} 个关联订单，无法删除'
            }), 400
        
        # 删除客户
        cur.execute('DELETE FROM customers WHERE id = ?', (customer_id,))
        
        if cur.rowcount == 0:
            conn.close()
            return jsonify({'success': False, 'message': '客户不存在'}), 404
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': '删除成功'})
        
    except Exception as e:
        logger.error(f"删除客户失败: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/customers/export', methods=['GET'])
@login_required
def export_customers():
    """导出客户数据到Excel"""
    try:
        import io
        from datetime import datetime
        
        search_name = request.args.get('name', '').strip()
        search_phone = request.args.get('phone', '').strip()
        
        conn = get_db()
        cur = conn.cursor()
        
        # 构建查询条件
        where_conditions = []
        params = []
        
        if search_name:
            where_conditions.append("customer_name LIKE ?")
            params.append(f'%{search_name}%')
        
        if search_phone:
            where_conditions.append("contact_phone LIKE ?")
            params.append(f'%{search_phone}%')
        
        where_clause = ' AND '.join(where_conditions) if where_conditions else '1=1'
        
        # 查询所有数据
        cur.execute(f'''
            SELECT customer_code, customer_name, contact_person, contact_phone,
                   address, remarks, create_time
            FROM customers 
            WHERE {where_clause}
            ORDER BY create_time DESC
        ''', params)
        
        rows = cur.fetchall()
        conn.close()
        
        # 创建Excel
        import pandas as pd
        df = pd.DataFrame(rows, columns=[
            '客户编号', '客户名称', '联系人', '联系电话', '收货地址', '备注', '创建时间'
        ])
        
        # 输出到内存
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='客户列表')
        
        output.seek(0)
        
        # 生成文件名
        filename = f'客户列表_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        from flask import send_file
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logger.error(f"导出客户数据失败: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500

# ========== 分拣标签管理 API ==========

def generate_label_code():
    """生成标签编号"""
    conn = get_db()
    cur = conn.cursor()
    today = datetime.now().strftime('%Y%m%d')
    prefix = f'FJ{today}'
    
    cur.execute('''
        SELECT label_code FROM picking_labels 
        WHERE label_code LIKE ? 
        ORDER BY label_code DESC LIMIT 1
    ''', (f'{prefix}%',))
    
    result = cur.fetchone()
    conn.close()
    
    if result:
        last_num = int(result[0][-4:])
        new_num = last_num + 1
    else:
        new_num = 1
    
    return f'{prefix}{new_num:04d}'


@app.route('/picking_labels')
@login_required
def picking_labels_page():
    """分拣标签管理页面"""
    return render_template('picking_labels.html')


@app.route('/api/picking_labels', methods=['GET'])
@login_required
def get_picking_labels():
    """获取分拣标签列表"""
    try:
        page = int(request.args.get('page', 1))
        page_size = int(request.args.get('page_size', 15))
        search_label_code = request.args.get('label_code', '').strip()
        search_order_code = request.args.get('order_code', '').strip()
        search_status = request.args.get('status', '').strip()
        
        conn = get_db()
        cur = conn.cursor()
        
        where_conditions = []
        params = []
        
        if search_label_code:
            where_conditions.append("label_code LIKE ?")
            params.append(f'%{search_label_code}%')
        
        if search_order_code:
            where_conditions.append("order_code LIKE ?")
            params.append(f'%{search_order_code}%')
        
        if search_status:
            where_conditions.append("label_status = ?")
            params.append(search_status)
        
        where_clause = ' AND '.join(where_conditions) if where_conditions else '1=1'
        
        cur.execute(f"SELECT COUNT(*) FROM picking_labels WHERE {where_clause}", params)
        total = cur.fetchone()[0]
        
        offset = (page - 1) * page_size
        cur.execute(f'''
            SELECT id, label_code, order_id, order_code, customer_name,
                   product_name, category, specification, quantity, unit,
                   delivery_date, label_status, print_count, remarks, create_time
            FROM picking_labels 
            WHERE {where_clause}
            ORDER BY create_time DESC
            LIMIT ? OFFSET ?
        ''', params + [page_size, offset])
        
        columns = [desc[0] for desc in cur.description]
        items = [dict(zip(columns, row)) for row in cur.fetchall()]
        
        conn.close()
        
        return jsonify({
        'success': True,
        'data': items,
        'total': total,
        'pagination': {
            'page': page,
            'page_size': page_size,
            'total_pages': (total + page_size - 1) // page_size
        }
    })
        
    except Exception as e:
        logger.error(f"获取标签列表失败: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/picking_labels/<int:label_id>', methods=['GET'])
@login_required
def get_picking_label(label_id):
    """获取单个标签详情"""
    try:
        conn = get_db()
        cur = conn.cursor()
        
        cur.execute('''
            SELECT id, label_code, order_id, order_code, customer_name,
                   product_name, category, specification, quantity, unit,
                   delivery_date, label_status, print_count, remarks
            FROM picking_labels WHERE id = ?
        ''', (label_id,))
        
        row = cur.fetchone()
        if not row:
            conn.close()
            return jsonify({'success': False, 'message': '标签不存在'}), 404
        
        columns = [desc[0] for desc in cur.description]
        label = dict(zip(columns, row))
        
        conn.close()
        
        return jsonify({'success': True, 'data': label})
        
    except Exception as e:
        logger.error(f"获取标签详情失败: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/picking_labels/generate', methods=['POST'])
@login_required
def generate_picking_labels():
    """从销售订单生成标签"""
    try:
        data = request.get_json()
        order_id = data.get('order_id')
        items = data.get('items', [])
        
        if not order_id:
            return jsonify({'success': False, 'message': '订单ID不能为空'}), 400
        
        if not items:
            return jsonify({'success': False, 'message': '请选择要生成标签的商品'}), 400
        
        conn = get_db()
        cur = conn.cursor()
        
        # 获取订单信息
        cur.execute('''
            SELECT order_code, customer_name, delivery_date
            FROM sales_orders WHERE id = ?
        ''', (order_id,))
        
        order = cur.fetchone()
        if not order:
            conn.close()
            return jsonify({'success': False, 'message': '订单不存在'}), 404
        
        order_code, customer_name, delivery_date = order
        
        # 为每个商品生成标签
        generated_count = 0
        for item in items:
            label_code = generate_label_code()
            
            cur.execute('''
                INSERT INTO picking_labels (
                    label_code, order_id, order_code, customer_name,
                    product_name, category, specification, quantity, unit,
                    delivery_date, label_status, create_user
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                label_code,
                order_id,
                order_code,
                customer_name,
                item['product_name'],
                item.get('category'),
                item.get('specification'),
                item['quantity'],
                item.get('unit', '件'),
                delivery_date,
                '待打印',
                session.get('user', 'system')
            ))
            generated_count += 1
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': f'成功生成 {generated_count} 个标签',
            'count': generated_count
        })
        
    except Exception as e:
        logger.error(f"生成标签失败: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/picking_labels/mark_printed', methods=['POST'])
@login_required
def mark_labels_printed():
    """标记标签为已打印"""
    try:
        data = request.get_json()
        label_ids = data.get('label_ids', [])
        
        if not label_ids:
            return jsonify({'success': False, 'message': '标签ID不能为空'}), 400
        
        conn = get_db()
        cur = conn.cursor()
        
        placeholders = ','.join(['?'] * len(label_ids))
        cur.execute(f'''
            UPDATE picking_labels 
            SET label_status = '已打印',
                print_count = print_count + 1
            WHERE id IN ({placeholders})
        ''', label_ids)
        
        conn.commit()
        updated = cur.rowcount
        conn.close()
        
        return jsonify({
            'success': True,
            'message': f'成功更新 {updated} 个标签',
            'count': updated
        })
        
    except Exception as e:
        logger.error(f"更新标签状态失败: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/picking_labels/<int:label_id>', methods=['DELETE'])
@login_required
def delete_picking_label(label_id):
    """删除标签"""
    try:
        conn = get_db()
        cur = conn.cursor()
        
        cur.execute('SELECT id FROM picking_labels WHERE id = ?', (label_id,))
        if not cur.fetchone():
            conn.close()
            return jsonify({'success': False, 'message': '标签不存在'}), 404
        
        cur.execute('DELETE FROM picking_labels WHERE id = ?', (label_id,))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': '删除成功'})
        
    except Exception as e:
        logger.error(f"删除标签失败: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500

# ========== 库存管理 API ==========

def generate_record_code(record_type):
    """生成入库/出库单号"""
    conn = get_db()
    cur = conn.cursor()
    today = datetime.now().strftime('%Y%m%d')
    prefix = f'{record_type}{today}'
    
    table_name = 'inbound_records' if record_type == 'RK' else 'outbound_records'
    
    cur.execute(f'''
        SELECT record_code FROM {table_name}
        WHERE record_code LIKE ? 
        ORDER BY record_code DESC LIMIT 1
    ''', (f'{prefix}%',))
    
    result = cur.fetchone()
    conn.close()
    
    if result:
        last_num = int(result[0][-4:])
        new_num = last_num + 1
    else:
        new_num = 1
    
    return f'{prefix}{new_num:04d}'


@app.route('/inventory', endpoint='inventory')
@login_required
def inventory_page():
    """库存管理页面"""
    return render_template('inventory.html')


@app.route('/api/inventory', methods=['GET'])
@login_required
def get_inventory():
    """获取库存列表"""
    try:
        page = int(request.args.get('page', 1))
        page_size = int(request.args.get('page_size', 15))
        search_product = request.args.get('product_name', '').strip()
        search_category = request.args.get('category', '').strip()
        search_stock_status = request.args.get('stock_status', '').strip()
        
        conn = get_db()
        cur = conn.cursor()
        
        where_conditions = []
        params = []
        
        if search_product:
            where_conditions.append("product_name LIKE ?")
            params.append(f'%{search_product}%')
        
        if search_category:
            where_conditions.append("category = ?")
            params.append(search_category)
        
        if search_stock_status:
            if search_stock_status == 'low':
                where_conditions.append("current_stock <= 0")
            elif search_stock_status == 'warning':
                where_conditions.append("current_stock > 0 AND current_stock <= safe_stock")
            elif search_stock_status == 'normal':
                where_conditions.append("current_stock > safe_stock")
        
        where_clause = ' AND '.join(where_conditions) if where_conditions else '1=1'
        
        cur.execute(f"SELECT COUNT(*) FROM inventory WHERE {where_clause}", params)
        total = cur.fetchone()[0]
        
        offset = (page - 1) * page_size
        cur.execute(f'''
            SELECT id, product_name, category, specification, unit,
                   current_stock, safe_stock, warehouse_location, remarks,
                   create_time, update_time
            FROM inventory 
            WHERE {where_clause}
            ORDER BY product_name
            LIMIT ? OFFSET ?
        ''', params + [page_size, offset])
        
        columns = [desc[0] for desc in cur.description]
        items = [dict(zip(columns, row)) for row in cur.fetchall()]
        
        conn.close()
        
        return jsonify({
        'success': True,
        'data': items,
        'total': total,
        'pagination': {
            'page': page,
            'page_size': page_size,
            'total_pages': (total + page_size - 1) // page_size
        }
    })
        
    except Exception as e:
        logger.error(f"获取库存列表失败: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/inventory/<int:product_id>', methods=['GET'])
@login_required
def get_inventory_item(product_id):
    """获取单个库存商品详情"""
    try:
        conn = get_db()
        cur = conn.cursor()
        
        cur.execute('''
            SELECT id, product_name, category, specification, unit,
                   current_stock, safe_stock, warehouse_location, remarks
            FROM inventory WHERE id = ?
        ''', (product_id,))
        
        row = cur.fetchone()
        if not row:
            conn.close()
            return jsonify({'success': False, 'message': '商品不存在'}), 404
        
        columns = [desc[0] for desc in cur.description]
        item = dict(zip(columns, row))
        
        conn.close()
        
        return jsonify({'success': True, 'data': item})
        
    except Exception as e:
        logger.error(f"获取库存详情失败: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/inventory', methods=['POST'])
@login_required
def add_inventory():
    """新增库存商品"""
    try:
        data = request.get_json()
        
        required_fields = ['product_name']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'success': False, 'message': f'{field}不能为空'}), 400
        
        conn = get_db()
        cur = conn.cursor()
        
        # 检查是否已存在
        cur.execute('''
            SELECT id FROM inventory 
            WHERE product_name = ? AND specification = ?
        ''', (data['product_name'], data.get('specification', '')))
        
        if cur.fetchone():
            conn.close()
            return jsonify({'success': False, 'message': '该商品已存在'}), 400
        
        cur.execute('''
            INSERT INTO inventory (
                product_name, category, specification, unit,
                current_stock, safe_stock, warehouse_location, remarks
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            data['product_name'],
            data.get('category', ''),
            data.get('specification', ''),
            data.get('unit', '件'),
            data.get('current_stock', 0),
            data.get('safe_stock', 0),
            data.get('warehouse_location', ''),
            data.get('remarks', '')
        ))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': '商品添加成功'})
        
    except Exception as e:
        logger.error(f"添加库存商品失败: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/inventory/<int:product_id>', methods=['PUT'])
@login_required
def update_inventory_item(product_id):
    """更新库存商品信息"""
    try:
        data = request.get_json()
        
        conn = get_db()
        cur = conn.cursor()
        
        cur.execute('SELECT id FROM inventory WHERE id = ?', (product_id,))
        if not cur.fetchone():
            conn.close()
            return jsonify({'success': False, 'message': '商品不存在'}), 404
        
        cur.execute('''
            UPDATE inventory SET
                category = ?,
                specification = ?,
                unit = ?,
                safe_stock = ?,
                warehouse_location = ?,
                remarks = ?,
                update_time = CURRENT_TIMESTAMP
            WHERE id = ?
        ''', (
            data.get('category', ''),
            data.get('specification', ''),
            data.get('unit', '件'),
            data.get('safe_stock', 0),
            data.get('warehouse_location', ''),
            data.get('remarks', ''),
            product_id
        ))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': '更新成功'})
        
    except Exception as e:
        logger.error(f"更新库存失败: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/inventory/statistics', methods=['GET'])
@login_required
def get_inventory_statistics():
    """获取库存统计数据"""
    try:
        conn = get_db()
        cur = conn.cursor()
        
        # 总商品数
        cur.execute('SELECT COUNT(*) FROM inventory')
        total_products = cur.fetchone()[0]
        
        # 低库存商品数
        cur.execute('SELECT COUNT(*) FROM inventory WHERE current_stock <= safe_stock')
        low_stock_count = cur.fetchone()[0]
        
        # 今日入库
        today = datetime.now().strftime('%Y-%m-%d')
        cur.execute('SELECT COUNT(*) FROM inbound_records WHERE inbound_date = ?', (today,))
        today_inbound = cur.fetchone()[0]
        
        # 今日出库
        cur.execute('SELECT COUNT(*) FROM outbound_records WHERE outbound_date = ?', (today,))
        today_outbound = cur.fetchone()[0]
        
        conn.close()
        
        return jsonify({
            'success': True,
            'data': {
                'total_products': total_products,
                'low_stock_count': low_stock_count,
                'today_inbound': today_inbound,
                'today_outbound': today_outbound
            }
        })
        
    except Exception as e:
        logger.error(f"获取统计数据失败: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/inventory/categories', methods=['GET'])
@login_required
def get_inventory_categories():
    """获取所有商品类别"""
    try:
        conn = get_db()
        cur = conn.cursor()
        
        cur.execute('SELECT DISTINCT category FROM inventory WHERE category IS NOT NULL AND category != "" ORDER BY category')
        categories = [row[0] for row in cur.fetchall()]
        
        conn.close()
        
        return jsonify({'success': True, 'data': categories})
        
    except Exception as e:
        logger.error(f"获取类别失败: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/inventory/inbound', methods=['POST'])
@login_required
def create_inbound():
    """创建入库记录"""
    try:
        data = request.get_json()
        
        required_fields = ['product_id', 'quantity', 'inbound_date']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'success': False, 'message': f'{field}不能为空'}), 400
        
        conn = get_db()
        cur = conn.cursor()
        
        # 获取商品信息
        cur.execute('''
            SELECT product_name, category, specification, unit, current_stock
            FROM inventory WHERE id = ?
        ''', (data['product_id'],))
        
        product = cur.fetchone()
        if not product:
            conn.close()
            return jsonify({'success': False, 'message': '商品不存在'}), 404
        
        product_name, category, specification, unit, current_stock = product
        
        # 生成入库单号
        record_code = generate_record_code('RK')
        
        # 插入入库记录
        cur.execute('''
            INSERT INTO inbound_records (
                record_code, inbound_type, product_name, category, specification,
                quantity, unit, purchase_order_code, supplier_name,
                inbound_date, operator, remarks
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            record_code,
            data.get('inbound_type', '采购入库'),
            product_name,
            category,
            specification,
            data['quantity'],
            unit,
            data.get('purchase_order_code', ''),
            data.get('supplier_name', ''),
            data['inbound_date'],
            session.get('user', 'system'),
            data.get('remarks', '')
        ))
        
        # 更新库存
        new_stock = current_stock + float(data['quantity'])
        cur.execute('''
            UPDATE inventory SET 
                current_stock = ?,
                update_time = CURRENT_TIMESTAMP
            WHERE id = ?
        ''', (new_stock, data['product_id']))
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': '入库成功',
            'record_code': record_code
        })
        
    except Exception as e:
        logger.error(f"入库失败: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/inventory/outbound', methods=['POST'])
@login_required
def create_outbound():
    """创建出库记录"""
    try:
        data = request.get_json()
        
        required_fields = ['product_id', 'quantity', 'outbound_date']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'success': False, 'message': f'{field}不能为空'}), 400
        
        conn = get_db()
        cur = conn.cursor()
        
        # 获取商品信息
        cur.execute('''
            SELECT product_name, category, specification, unit, current_stock
            FROM inventory WHERE id = ?
        ''', (data['product_id'],))
        
        product = cur.fetchone()
        if not product:
            conn.close()
            return jsonify({'success': False, 'message': '商品不存在'}), 404
        
        product_name, category, specification, unit, current_stock = product
        
        # 检查库存是否足够
        quantity = float(data['quantity'])
        if current_stock < quantity:
            conn.close()
            return jsonify({'success': False, 'message': f'库存不足，当前库存：{current_stock}'}), 400
        
        # 生成出库单号
        record_code = generate_record_code('CK')
        
        # 插入出库记录
        cur.execute('''
            INSERT INTO outbound_records (
                record_code, outbound_type, product_name, category, specification,
                quantity, unit, sales_order_code, customer_name,
                outbound_date, operator, remarks
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            record_code,
            data.get('outbound_type', '销售出库'),
            product_name,
            category,
            specification,
            quantity,
            unit,
            data.get('sales_order_code', ''),
            data.get('customer_name', ''),
            data['outbound_date'],
            session.get('user', 'system'),
            data.get('remarks', '')
        ))
        
        # 更新库存
        new_stock = current_stock - quantity
        cur.execute('''
            UPDATE inventory SET 
                current_stock = ?,
                update_time = CURRENT_TIMESTAMP
            WHERE id = ?
        ''', (new_stock, data['product_id']))
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': '出库成功',
            'record_code': record_code
        })
        
    except Exception as e:
        logger.error(f"出库失败: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/inventory/inbound_records', methods=['GET'])
@login_required
def get_inbound_records():
    """获取入库记录"""
    try:
        page = int(request.args.get('page', 1))
        page_size = int(request.args.get('page_size', 15))
        
        conn = get_db()
        cur = conn.cursor()
        
        cur.execute('SELECT COUNT(*) FROM inbound_records')
        total = cur.fetchone()[0]
        
        offset = (page - 1) * page_size
        cur.execute('''
            SELECT id, record_code, inbound_type, product_name, specification,
                   quantity, unit, supplier_name, inbound_date, operator, remarks
            FROM inbound_records
            ORDER BY inbound_date DESC, create_time DESC
            LIMIT ? OFFSET ?
        ''', [page_size, offset])
        
        columns = [desc[0] for desc in cur.description]
        items = [dict(zip(columns, row)) for row in cur.fetchall()]
        
        conn.close()
        
        return jsonify({
            'success': True,
            'data': {
                'items': items,
                'total': total,
                'page': page,
                'page_size': page_size
            }
        })
        
    except Exception as e:
        logger.error(f"获取入库记录失败: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/inventory/outbound_records', methods=['GET'])
@login_required
def get_outbound_records():
    """获取出库记录"""
    try:
        page = int(request.args.get('page', 1))
        page_size = int(request.args.get('page_size', 15))
        
        conn = get_db()
        cur = conn.cursor()
        
        cur.execute('SELECT COUNT(*) FROM outbound_records')
        total = cur.fetchone()[0]
        
        offset = (page - 1) * page_size
        cur.execute('''
            SELECT id, record_code, outbound_type, product_name, specification,
                   quantity, unit, customer_name, outbound_date, operator, remarks
            FROM outbound_records
            ORDER BY outbound_date DESC, create_time DESC
            LIMIT ? OFFSET ?
        ''', [page_size, offset])
        
        columns = [desc[0] for desc in cur.description]
        items = [dict(zip(columns, row)) for row in cur.fetchall()]
        
        conn.close()
        
        return jsonify({
            'success': True,
            'data': {
                'items': items,
                'total': total,
                'page': page,
                'page_size': page_size
            }
        })
        
    except Exception as e:
        logger.error(f"获取出库记录失败: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500

if __name__ == '__main__':
        app.run(debug=True, host='0.0.0.0', port=5000)